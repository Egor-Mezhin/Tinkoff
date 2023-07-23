from tkinter import *
from tkinter import ttk
from openpyxl import Workbook
import pyperclip
import sqlite3
import ctypes
import math
import sys
import os

color = {
    "Активная":"#F7C815",
    "Неактивная":"#deb416",
    "Сохранить":"#228b22",
    "Изменить":"#FAD074",
    "Удалить":"#ff2400",
    "Активная категория":"#EC9704",
    "Текст":"#fff",
    "Обводка_сборочных":"#F7E8E3",
}

font = {
    "Большой":('Helvetica', 12),
    "Средний":('Helvetica', 10),
}

conn = sqlite3.connect('orders.db')
cur = conn.cursor()

if sys.platform == 'win32' or sys.platform.startswith('linux'):
    ctrl = "<Control-KeyPress>"
elif sys.platform == "darwin":
    ctrl = "<Command-KeyPress>"

def get_path(event):
    print(event.data)

def error_two_open():
    error = Tk()
    error.title("TCRM+")
    error.geometry("400x200")
    error.attributes("-topmost",True)
    error.iconbitmap("icon.ico")

    error_label = Label(error, text = "ОШИБКА:\nНе удалось отправить запрос на базу данных.\nЗакройте все окошки приложения и откройте только одно.\nТогда ошибка не повторится")
    error_label.pack()
    root.destroy()

def is_ru_lang_keyboard():
    """
    Функция для работы функции keys. Для работы блокнота в русской раскладке
    """
    if sys.platform == 'win32':
        u = ctypes.windll.LoadLibrary("user32.dll")
        pf = getattr(u, "GetKeyboardLayout")
        return hex(pf(0)) == '0x4190419'
    elif sys.platform == "darwin":
        u = ctypes.cdll.LoadLibrary("/System/Library/Frameworks/ApplicationServices.framework/Versions/A/Frameworks/HIServices.framework/Versions/A/HIServices")
        pf = getattr(u, "TISGetInputSourceProperty")
        property_name = ctypes.c_void_p.in_dll(u, "kTISPropertyUnicodeKeyLayoutData")
        handle = pf(ctypes.c_void_p.in_dll(u, "kTISPropertyInputSourceClass"), property_name)
        return bool(handle)
    elif sys.platform.startswith('linux'):
        u = ctypes.cdll.LoadLibrary('libX11.so')
        display = u.XOpenDisplay(None)
        if display == 0:
            return False
        keysym = u.XStringToKeysym('F2')
        keycode = u.XKeysymToKeycode(display, keysym)
        return bool(keycode)
    return False 

def keys(event):
    if is_ru_lang_keyboard():
        if event.keycode==86: 
            event.widget.event_generate("<<Paste>>")

        elif event.keycode==67: 
            pyperclip.copy(event.widget.selection_get())

        elif event.keycode==88:  
            pyperclip.copy(event.widget.selection_get())
            event.widget.delete("sel.first", "sel.last")
            
        elif event.keycode==65535: 
            event.widget.event_generate("<<Clear>>")

        elif event.keycode==65: 
            event.widget.event_generate("<<SelectAll>>")

        elif event.keycode == 90:
            event.widget.edit_undo()
        
        elif event.keycode == 8:
            event.widget.delete("insert-1c wordstart", "insert")
            
    elif event.keycode==67: 
        pyperclip.copy(event.widget.selection_get())
        return "break"
    
    elif event.keycode==88:  
        pyperclip.copy(event.widget.selection_get())
        event.widget.delete("sel.first", "sel.last")
        return "break"
    
    elif event.keycode == 8:
        event.widget.delete("insert-1c wordstart", "insert")
        

def copy(val, buffer = None):
    """
    Копирование данных в буфер обмена
    
    val - Копируемый текст
    buffer - Текстовое поле где отражен скопированный текст

    pyperclip.copy - копирует val в буфер обмена
    """

    pyperclip.copy(val.format(main.aut_text_entry.get()))

    if buffer != None:
        buffer.delete("1.0", END)
        buffer.insert("2.0", val.format(main.aut_text_entry.get()))

root = Tk()
root.title("TCRM+")
root.geometry("265x673")
root.minsize(width=200, height=640)
root.maxsize(width=300, height=700)
root.attributes("-topmost",True)

my_file = os.path.abspath(sys.argv[0]) 
my_dir = os.path.dirname(my_file)
icon_file = os.path.join(my_dir, "icon.ico")

if os.path.exists(icon_file):
    root.iconbitmap(icon_file)

class main:

    """
    Основная панель, содержит: 
    Поле вставки из буфера, 
    счетчик,
    Вкладки.

    /////// Функции ///////
    
    f_counter_minus, f_counter_plus, f_counter_null - Функции контроля счетчика +, -, 0

    """

    def past():

        """
        Вставляет в текстовое поле содержимое буфера обмена
        
        empty - Поле в которое нужно вставить текст

        empty_str - Содержание буфера обмена
        empty_new - Новый текст если в конце empty_str есть пробел
        """

        empty_str = pyperclip.paste()
        empty_new = str()
        if empty_str[-1] == " " and len(empty_str) > 0:
            empty_new = empty_str[0:len(empty_str)-1]
        else:
            empty_new = empty_str
        
        main.aut_text_entry.delete(0, END)
        main.aut_text_entry.insert(0, empty_new)


    def f_counter_minus():
        
        """
        Убавляет значение 1 со счетчика

        cur.execute - Уменьшает счетчик в БД
        new_empty - новое значение счетчика
        """

        if main.counter["text"] > "0":

            cur.execute("""
            UPDATE counter 
            SET count = count - 1""")
            conn.commit()

            new_empty = str(int(main.counter["text"]) - 1)
            main.counter["text"] = new_empty


    def f_counter_plus():

        """
        Добавляет к счетчику значение 1

        cur.execute - Увеличивает счетчик в БД
        new_empty - новое значение счетчика
        """

        cur.execute("""
        UPDATE counter 
        SET count = count + 1""")
        conn.commit()

        new_empty = str(int(main.counter["text"]) + 1)
        main.counter["text"] = new_empty


    def f_counter_null():
        
        """
        Обнуляет счетчик

        cur.execute - Обнуляет счетчик в БД
        """

        cur.execute("""
        UPDATE counter 
        SET count = 0""")
        conn.commit()

        main.counter["text"] = "0"

    """  
    /////// Фреймы /////// 

    frame - Основной фрейм растянутый на все поле
    frame_up - Фрейм шапки приложения
    frame_up_aut_text - Фрейм для вставки
    frame_up_counter - Фрейм счетчика
    frame_bot - Фрейм Вкладок - растянут на всю доступную высоту
    frame_Ready, frame_assembling, frame_settings, frame_book - Фреймы для закладок

    /////// Элементы /////// 

    aut_text_btn - Кнопка вставки в поле aut_text_entry
    aut_text_entry - Текстовое поле. С него копируют текст для скрипта(обычно имя)
    counter_plus, counter_minus, counter_null - Кнопки контроля счетчика
    counter - Поле ввода счетчика
    notebook - Класс вкладок

    /////// Переменные /////// 

    count - значение счетчика из БД
    """
    frame = Frame(root)
    frame.place(rely=0, relheight=1, relwidth=1)

    frame_aut_text = Frame(frame)
    frame_aut_text.pack(fill=X)

    frame_counter = Frame(frame)
    frame_counter.pack(fill=X)

    frame_bot = Frame(frame)
    frame_bot.pack(fill=X)

    aut_text_btn = Button(frame_aut_text, text = "Вставить", width=10, command= lambda: main.past())
    aut_text_btn.grid(column=0, row = 0)

    aut_text_entry = Entry(frame_aut_text, width=100)
    aut_text_entry.grid(column=1, row = 0)
    aut_text_entry.bind(ctrl, keys)

    counter_plus = Button(frame_counter, text = "+", width=10, command = f_counter_plus)
    counter_plus.grid(column=0, row = 0, sticky="nswe")

    counter_minus = Button(frame_counter, text = "-", width=10, command = f_counter_minus)
    counter_minus.grid(column=1, row = 0, sticky="nswe")

    counter_null = Button(frame_counter, text = "0", width=1, command = f_counter_null)
    counter_null.grid(column=2, row = 0)

    cur.execute("SELECT count FROM counter")
    count = str(cur.fetchone()[0])
    counter = Label(frame_counter, text= count, font = (font["Большой"], 12))
    counter.grid(column=3, row = 0)

    notebook = ttk.Notebook(frame_bot)
    notebook.pack(expand=True, fill=BOTH)

    frame_assembling = ttk.Frame(notebook)
    frame_Ready = ttk.Frame(notebook)
    frame_settings = ttk.Frame(notebook)
    frame_book = ttk.Frame(notebook)


    frame_assembling.pack(fill=BOTH, expand=True)
    frame_Ready.pack(fill=BOTH, expand=True)
    frame_settings.pack(fill=BOTH, expand=True)
    frame_book.pack(fill=BOTH, expand=True)

    notebook.add(frame_assembling, text="СБОРКА")
    notebook.add(frame_Ready, text="ГОТОВЫЕ")
    notebook.add(frame_settings, text="НАСТРОЙКА")
    notebook.add(frame_book, text="Б")


class Ready:
    """
    Вкладка готовых скриптов. Содержит вкладки категорий скриптов и кнопки для копирования скриптов
    
    /////// Функции ///////
    
    ready_next_page, ready_first_page, ready_back_page - Управление страницами скриптов +1, -1, 1
    ready_swap_category - Переключение категорий при нажатии на кнопку категории
    """


    def ready_next_page():
        """
        Переключение страниц скриптов вперед для вкладки готовых скриптов

        val_list - Новый список скриптов
        key - Титульник скрипта
        val - Текст скрипта

        Цикл 1: - Переименование текста и закрепленных за кнопкой скриптов
            Исключение IndexError - Если в списке закончились значения то кнопка неактивна
        """

        c = Ready

        if c.page != c.len_ready_text:
            c.page += 1
            c.page_list += 10

            c.center_btn.configure(text = f"{c.page}/{c.len_ready_text}")
            
            cur.execute(f"""
                        SELECT ready_text.title, ready_text.text
                        FROM ready_text
                        JOIN ready_category ON ready_category.id = ready_text.id_category
                        WHERE ready_category.title = "{c.category}"
                        LIMIT 10 OFFSET {c.page_list};
                        """)
            val_list = cur.fetchall()

            for i_num, i_val in enumerate(c.btn_script_ready_list): #Цикл 1
                try:

                    key = val_list[i_num][0]
                    val = val_list[i_num][1]
                    
                    i_val.configure(text = key, state='normal', command = lambda val=val: copy(val))

                except IndexError: #Исключение IndexError
                    i_val.configure(text = "", state='disabled', bg = color["Неактивная"])


    def ready_back_page():
        """
        Переключение страниц скриптов назад для вкладки готовых скриптов

        val_list - Новый список скриптов
        key - Титульник скрипта
        val - Текст скрипта

        Цикл 1: - Переименование текста и закрепленных за кнопкой скриптов
        """

        c = Ready

        if c.page != 1:
            c.page -= 1
            c.page_list -= 10

            c.center_btn.configure(text = f"{c.page}/{c.len_ready_text}")
            
            cur.execute(f"""
                        SELECT ready_text.title, ready_text.text
                        FROM ready_text
                        JOIN ready_category ON ready_category.id = ready_text.id_category
                        WHERE ready_category.title = "{c.category}"
                        LIMIT 10 OFFSET {c.page_list};
                        """)
            val_list = cur.fetchall()

            for i_num, i_val in enumerate(c.btn_script_ready_list): # Цикл 1
                try:
                    key = val_list[i_num][0]
                    val = val_list[i_num][1]

                    i_val.configure(text = key, bg = color["Активная"], state='normal', command = lambda val=val: copy(val))
                except IndexError:
                    i_val.configure(text = "", state='disabled', bg = color["Неактивная"])


    def ready_first_page():
        """
        Переключение страниц скриптов на первую для вкладки готовых скриптов
        
        val_list - Новый список скриптов
        key - Титульник скрипта
        val - Текст скрипта

        Цикл 1: - Переименование текста и закреплных за кнопкой скриптов
        
        """

        c = Ready
        
        c.page = 1
        c.page_list = 0

        c.center_btn.configure(text = f"{c.page}/{c.len_ready_text}")
        
        cur.execute(f"""
                    SELECT ready_text.title, ready_text.text
                    FROM ready_text
                    JOIN ready_category ON ready_category.id = ready_text.id_category
                    WHERE ready_category.title = "{c.category}"
                    LIMIT 10 OFFSET {c.page_list};
                    """)
        val_list = cur.fetchall()

        for i_num, i_val in enumerate(c.btn_script_ready_list): # Цикл 1
            try:
                val = val_list[i_num][1]
                i_val.configure(text = val_list[i_num][0], state='normal', bg = color["Активная"], command = lambda val=val: copy(val))

            except IndexError:
                i_val.configure(text = "", state='disabled', bg = color["Неактивная"])


    def ready_swap_category(category):
        """
        Переключение категорий скриптов для вкладки готовых скриптов

        ready_text - Новый список скриптов
        len_ready_text - новое максимальное значение для колличества страниц

        Цикл 1 - Перекраска кнопок категорий.
        Цикл 2 - Переименование кнопок из ready_text 
            Исключение IndexError - Если скрипты закончились в списке то кнопка неактивна
        """
        c = Ready
        c.category = category
        c.page = 1
        c.page_list = 0

        cur.execute(f"""
        SELECT ready_text.title, ready_text.text 
        FROM ready_text
        JOIN ready_category ON 
        ready_category.id = ready_text.id_category
        WHERE ready_category.title = "{c.category}"
        ORDER BY ready_text.id;
        """)
        ready_text = cur.fetchall()
        
        len_ready_text = math.ceil(len(ready_text) / 10)
        if len_ready_text == 0:
            len_ready_text = 1
            
        c.len_ready_text = len_ready_text

        for i_val in c.category_list: # Цикл 1
            if i_val["text"] == category:
                i_val.configure(state='disabled', bg = color["Активная категория"]) 
            else:
                i_val.configure(state='normal', bg = color["Активная"])

        for i_val in range(10): # Цикл 2
            try:
                key = ready_text[i_val][0]
                val = ready_text[i_val][1]

                c.btn_script_ready_list[i_val].configure(text = key, state='normal', bg = color["Активная"], command= lambda val=val: copy(val))

            except IndexError: # Исключение IndexError
                c.btn_script_ready_list[i_val].configure(text = "", state='disabled', bg = color["Неактивная"]) 
        
        c.center_btn.configure(text = f"1/{len_ready_text}")

    """ 
    /////// Фреймы /////// 

    Canvas_top - фрейм для кнопок категорий и скроллбара
    Сanvas_frame - Фрейм для кнопок категорий
    Canvas_top_scrollbar - Фрейм для скроллбара
    Frame_bottom - Нижний фрейм для стрелок и кнопок
    Frame_bottom_arrow - Фрейм для стрелок
    Frame_bottom_text - Фрейм для кнопок

    /////// Элементы /////// 
    
    scrollbar - скроллбар для прокрутки 
    btn_readys - Кнопка категории
    btn_script_ready - Кнопка скрипта категори
    left_btn - Кнопка страница скриптов назад
    center_btn - Кнопка первая страница скриптов 
    right_btn - Кнопка страница скриптов вперед

    /////// Переменные /////// 

    ready_text - Список всех значений с выбранной категории
    ready_category - список категорий
    category - Титульник первой категории в спике категорий
    with_list - Размер Canvas_top
    page - Страница скриптов
    page_list - Начальный элемент страницы скриптов
    len_ready_text - Максимальное колличество страниц
    category_list - Лист для категорий скриптов
    btn_script_ready_list - Лист для кнопок категорий скриптов

    /////// Циклы /////// 
    Цикл 1 - Задать название кнопкам категорий btn_readys из списка категорий ready_category
        val - Титульник категорий
        width_btn - длинна кнопки
    Цикл 2 - Создание 10 кнопок и генерация названий кнопок 
        key - Титульник кнопки
        val - Текст кнопки
        IndexError - Если в списке закончились названия кнопок то кнопка неактивна


    """

    Canvas_top = Canvas(main.frame_Ready, height=25)
    Canvas_top.pack(anchor=NW, fill=X)

    Сanvas_frame = Frame(Canvas_top, height=25)

    Canvas_top_scrollbar = Frame(main.frame_Ready, height=10)
    Canvas_top_scrollbar.pack(anchor=NW, fill=X)

    scrollbar = Scrollbar(Canvas_top_scrollbar, orient=HORIZONTAL, command=Canvas_top.xview)
    scrollbar.pack(expand=True, fill=BOTH)
    

    Frame_bottom = Frame(main.frame_Ready)
    Frame_bottom.pack(expand=True, fill=BOTH)

    Frame_bottom_arrow = Frame(Frame_bottom)
    Frame_bottom_arrow.pack(anchor=NW, fill=X)
    
    Frame_bottom_text = Frame(Frame_bottom)
    Frame_bottom_text.pack(anchor=NW, expand=True, fill=BOTH)
    
    cur.execute("""SELECT title FROM ready_category ORDER BY num;""")
    ready_category = cur.fetchall()

    category = ready_category[0][0]
    
    cur.execute(f"""
    SELECT ready_text.title, ready_text.text 
    FROM ready_text
    JOIN ready_category ON 
    ready_category.id = ready_text.id_category
    WHERE ready_category.id = 1
    ORDER BY ready_text.id;
    """)
    ready_text = cur.fetchall()

    with_list = int()

    page = 1
    page_list = 0
    
    len_ready_text = math.ceil(len(ready_text) / 10)
    if len_ready_text == 0:
        len_ready_text = 1
    
    category_list = list()
    btn_script_ready_list = list()

    for i_num, i_val in enumerate(ready_category): # Цикл 1
        
        val = i_val[0]
        
        width_btn = int(len(str(val)))
        with_list += width_btn

        btn_readys = Button(Сanvas_frame, text = val, width = width_btn,  bg = color["Активная"], font = font["Большой"], command = lambda val=val: Ready.ready_swap_category(val)) 
        btn_readys.pack(side=LEFT, expand=True, fill="both")
        category_list.append(btn_readys)

    category_list[0].configure(state='disabled', bg = color["Активная категория"])
    btn_readys_end = Button(Сanvas_frame, text = "             ", state='disabled', bg = color["Активная"]) 
    btn_readys_end.pack(side=LEFT, expand=True, fill="both")
    category_list.append(btn_readys_end)

    left_btn = Button(Frame_bottom_arrow, text = "<<<<<", bg = color["Активная"], font = font["Средний"], command = ready_back_page)
    left_btn.pack(side = LEFT, expand=True, fill="both")

    center_btn = Button(Frame_bottom_arrow, text = f"{page}/{len_ready_text}", width = 4, bg = color["Активная"], font = font["Средний"], command = ready_first_page)
    center_btn.pack(side = LEFT)
    
    right_btn = Button(Frame_bottom_arrow, text = ">>>>>", bg = color["Активная"], font = font["Средний"], command = ready_next_page)
    right_btn.pack(side = LEFT, expand=True, fill="both")

    for i_val in range(10): # Цикл 2
        
        try:
            key = ready_text[i_val][0]
            val = ready_text[i_val][1]

            btn_script_readys = Button(Frame_bottom_text, text = key, bg = color["Активная"], font = font["Большой"], command= lambda val=val: copy(val)) 

        except IndexError:
            btn_script_readys = Button(Frame_bottom_text, text = "", state='disabled', bg = color["Неактивная"], font = font["Большой"]) 
        finally:
            btn_script_readys.pack(fill = X, pady = 10)
            btn_script_ready_list.append(btn_script_readys)
    
    Canvas_top.create_window(0, 0, anchor=NW, window=Сanvas_frame, height=25)
    Canvas_top.config(scrollregion=(0, 0, with_list*12, 0))

class Assembling:
    """
    Вкладка сборочных скриптов.

    /////// Функции /////// 
    add_copy - Функция кнопки добавления скрипта

    assembling_next_page, assembling_back_page, assembling_first_page - Управление страницами скриптов

    """

    
    def add_copy(val, buffer):
        """
        Добавление к копированным данным в буфер обмена
        
        val - Копируемый текст
        buffer - Текстовое поле где отражен скопированный текст

        text - Соединяет содержание буфера обмена с val
        pyperclip.copy - копирует text в буфер обмена
        buffer.insert - Вставляет в текстовое поле val
        """
            
        text = pyperclip.paste() + val.format(main.aut_text_entry.get())

        pyperclip.copy(text)

        buffer.insert(END, val.format(main.aut_text_entry.get()))


    def assembling_next_page():
        """
        Переключение страниц скриптов вперед для вкладки сборочных скриптов
        
        val_list - Новый список скриптов

        Цикл 1 - Переименование текста скриптов
        Цикл 2 - Переименование закрепленных скриптов на кнопках копирования
        Цикл 3 - Переименование закрепленных скриптов на кнопках добавления
            IndexError - Если скрипты закончились в списке то кнопки неактивны текста нет
        """
        c = Assembling
        if c.page != c.len_assembling_list:
            c.page += 1
            c.page_list += 14
            c.center_btn.configure(text = f"{c.page}/{c.len_assembling_list}")

            cur.execute(f"SELECT title, text FROM assembling_list LIMIT 13 OFFSET {c.page_list - 1};")
            val_list = cur.fetchall()

            for i_num, i_val in enumerate(c.label_list): # Цикл 1 
                try:
                    i_val.configure(text = val_list[i_num][0])
                    
                except IndexError:
                    i_val.configure(text = "")
            
            for i_num, i_val in enumerate(c.btn_assembling_list): # Цикл 2
                try:
                    val = val_list[i_num][1]
                    i_val.configure(bg = color["Активная"], command = lambda val=val: copy(val, c.buffer_assembling))

                except IndexError:
                    i_val.configure(state='disabled', bg = color["Неактивная"])

            for i_num, i_val in enumerate(c.add_btn_assembling_list): # Цикл 3
                try:
                    val = val_list[i_num][1]
                    i_val.configure(bg = color["Активная"], command = lambda val=val: c.add_copy(val, c.buffer_assembling))

                except IndexError:
                    i_val.configure(state='disabled', bg = color["Неактивная"])
        

    def assembling_back_page():
        """
        Переключение страниц скриптов назад для вкладки сборочных скриптов

        val_list - Новый список скриптов

        Цикл 1 - Переименование текста скриптов
        Цикл 2 - Переименование закрепленных скриптов на кнопках копирования
        Цикл 3 - Переименование закрепленных скриптов на кнопках добавления
        """
        c = Assembling
        if c.page != 1:
            c.page -= 1
            c.page_list -= 14
            c.center_btn.configure(text = f"{c.page}/{c.len_assembling_list}")

            cur.execute(f"SELECT title, text FROM assembling_list LIMIT 13 OFFSET {c.page_list - 1};")
            val_list = cur.fetchall()

            for i_num, i_val in enumerate(c.label_list): # Цикл 1
                try:
                    i_val.configure(text = val_list[i_num][0])
                except IndexError:
                    i_val.configure(text='')

            for i_num, i_val in enumerate(c.btn_assembling_list): # Цикл 2
                try:
                    val = val_list[i_num][1]
                    i_val.configure(state='normal', bg = color["Активная"], command = lambda val=val: copy(val, c.buffer_assembling))
                except IndexError:
                    i_val.configure(state='disabled', bg = color["Неактивная"])

            for i_num, i_val in enumerate(c.add_btn_assembling_list): # Цикл 3
                try:
                    val = val_list[i_num][1]
                    i_val.configure(state='normal', bg = color["Активная"], command = lambda val=val: c.add_copy(val, c.buffer_assembling))
                except IndexError:
                    i_val.configure(state='disabled', bg = color["Неактивная"])

    def assembling_first_page():
        """
        Переключение страниц скриптов на первую для вкладки сборочных скриптов

        val_list - Новый список скриптов

        Цикл 1 - Переименование текста скриптов
        Цикл 2 - Переименование закрепленных скриптов на кнопках копирования
        Цикл 3 - Переименование закрепленных скриптов на кнопках добавления

        """
        c = Assembling
        c.page = 1
        c.page_list = 0
        c.center_btn.configure(text = f"{c.page}/{c.len_assembling_list}")

        cur.execute(f"SELECT title, text FROM assembling_list LIMIT 13 OFFSET {c.page_list - 1};")
        val_list = cur.fetchall()

        for i_num, i_val in enumerate(c.label_list): # Цикл 1
            try:
                i_val.configure(text = val_list[i_num][0])
            except IndexError:
                i_val.configure(text = "")
        
        for i_num, i_val in enumerate(c.btn_assembling_list): # Цикл 2
            try:
                val = val_list[i_num][1]
                i_val.configure(state='normal', bg = color["Активная"], command = lambda val=val: copy(val, c.buffer_assembling))
            except IndexError:
                i_val.configure(state='disabled', bg = color["Неактивная"])

        for i_num, i_val in enumerate(c.add_btn_assembling_list): # Цикл 3
            try:
                val = val_list[i_num][1]
                i_val.configure(state='normal', bg = color["Активная"], command = lambda val=val: c.add_copy(val, c.buffer_assembling))
            except IndexError:
                i_val.configure(state='disabled', bg = color["Неактивная"])
    """ 
    /////// Фреймы /////// 

    frame_assembling_left - Фрейм для кнопок копировать и добавить
    frame_assembling_right - Фрейм для титульников для кнопок копирования
    frame_assembling_bot - Фрейм для кнопок переключения и текстового поля буфера
    frame_assembling_bot_arrow - Фрейм для кнопок переключения
    frame_assembling_bot_text - Фрейм для текстового поля буфера
    frame_lb - фрейм для текста
    
    /////// Элементы /////// 

    btn_assembling - Кнопка копировать
    add_btn_assembling - кнопка добавить к скопированному
    lb - Текстовое поле для титульника кнопки
    left_btn - Кнопка страница скриптов назад
    center_btn - Кнопка первая страница скриптов 
    right_btn - Кнопка страница скриптов вперед
    c.buffer_assembling - текстовое поле для отображения скопированного

    /////// Переменные /////// 

    page - Страница скриптов
    page_list - Номер начального скрипта  
    assembling_list - список титульников и текстов сборочных скриптов
    btn_assembling_list - Список кнопок для копирования
    add_btn_assembling_list - Список кнопок для добавления к скопированному
    label_list - Список текстов
    count - счетчик для столбцов
    

    /////// Циклы /////// 

    Цикл 1 - Создает 14 фреймов для кнопко и присваиваем им текст и титульники
        IndexError - Если в списке нет значений то кнопки неактивны
        key - титульник кнопки
        val - Текст кнопки
        lb - титульник для кнопок копирования
    """
    frame_assembling_bot_arrow = Frame(main.frame_assembling, height=25)
    frame_assembling_bot_arrow.pack(fill = X)

    frame_assembling_script = Frame(main.frame_assembling, height=400)
    frame_assembling_script.pack(fill = X)

    frame_assembling_bot_text = Frame(main.frame_assembling, height=300)
    frame_assembling_bot_text.pack(fill = X)


    page = 1
    page_list = 0

    cur.execute("SELECT title, text FROM assembling_list;")
    assembling_list = cur.fetchall()

    len_assembling_list = math.ceil(len(assembling_list) / 14)
    if len_assembling_list == 0:
        len_assembling_list = 1
        
    btn_assembling_list = []
    add_btn_assembling_list = []
    label_list = []

    left_btn = Button(frame_assembling_bot_arrow, text = "<<<<<", bg = color["Активная"], font = font["Средний"], command = assembling_back_page) 
    left_btn.pack(side=LEFT, expand=True, fill="both")

    center_btn = Button(frame_assembling_bot_arrow, text = f"{page}/{len_assembling_list}", width = 5,  bg = color["Активная"], font = font["Средний"], command = assembling_first_page) 
    center_btn.pack(side=LEFT)

    right_btn = Button(frame_assembling_bot_arrow, text = ">>>>>", bg = color["Активная"], font = font["Средний"], command = assembling_next_page) 
    right_btn.pack(side=LEFT, expand=True, fill="both")

    for i_val in range(13): # Цикл 1
        
        frame_assembling_script_panel = Frame(frame_assembling_script, height=30, highlightthickness = 1, highlightbackground=color["Обводка_сборочных"], relief=SOLID)
        frame_assembling_script_panel.pack(fill = X)

        frame_assembling_script_panel_btn = Frame(frame_assembling_script_panel, height=30)
        frame_assembling_script_panel_btn.pack(side=LEFT)

        frame_assembling_script_panel_script = Frame(frame_assembling_script_panel, height=30)
        frame_assembling_script_panel_script.pack(side=LEFT)

        try:
            key = assembling_list[i_val][0]
            val = assembling_list[i_val][1]

            btn_assembling = Button(frame_assembling_script_panel_btn, text = "С", width = 3, bg = color["Активная"], font = font["Большой"], command = lambda val=val: copy(val, Assembling.buffer_assembling))
            add_btn_assembling = Button(frame_assembling_script_panel_btn, text = "+", width = 3, bg = color["Активная"], font = font["Большой"], command = lambda val=val: Assembling.add_copy(val, Assembling.buffer_assembling))
            lb = Label(frame_assembling_script_panel_script,  text = key, font = font["Большой"])

        except IndexError:

            btn_assembling = Button(frame_assembling_script_panel_btn, text = "С", width = 3, state='disabled', font = font["Большой"], bg = color["Неактивная"])
            add_btn_assembling = Button(frame_assembling_script_panel_btn, text = "+", width = 3, state='disabled', font = font["Большой"], bg = color["Неактивная"])
            lb = Label(frame_assembling_script_panel_script,  text = "", font = font["Большой"])
           
        finally:
            btn_assembling.pack(side=LEFT)
            btn_assembling_list.append(btn_assembling)

            add_btn_assembling.pack(side=LEFT)
            add_btn_assembling_list.append(add_btn_assembling)

            lb.grid(column = 0, row = 0)
            label_list.append(lb)


    buffer_assembling = Text(frame_assembling_bot_text, font = font["Большой"],wrap=WORD)
    buffer_assembling.pack(fill=BOTH)
    buffer_assembling.bind("<Key>", lambda e: "break") 

class Settings:
    """
    Вкладка настроек
    
    Имеет подвкладки:
    frame_assembling - Настройка сборочных скриптов
    frame_Ready - Настройка Готовых скриптов
    frame_general - Общие настройки

    """
    notebook = ttk.Notebook(main.frame_settings)
    notebook.pack(expand=True, fill=BOTH)

    # создаем пару фреймвов
    frame_assembling = ttk.Frame(notebook)
    frame_Ready = ttk.Frame(notebook)
    frame_general = ttk.Frame(notebook)

    frame_assembling.pack(fill=BOTH, expand=True)
    frame_Ready.pack(fill=BOTH, expand=True)
    frame_general.pack(fill=BOTH, expand=True)


    # добавляем фреймы в качестве вкладок
    notebook.add(frame_assembling, text="Сборка")
    notebook.add(frame_Ready, text="Готовые")
    notebook.add(frame_general, text="Общие")


class Settings_Assembling:
    """
    Вкладка настроек сборочных скриптов

    /////// Функции /////// 

    switching_assembling_combobox - Евент при переключении значений комбобокса choice_val_menu
    assembling_delete - Удаляет выбранный скрипт
    assembling_save - Добавляет новый скрипт или изменяет старый
    """
    def commit_script(status):
        c = Settings_Assembling

        conn.commit()
        cur.execute("SELECT id, title, text FROM assembling_list;") # SQL 3
        c.val_list = cur.fetchall()

        num_list = [i_val[0] for i_val in c.val_list]
        num_list.append(len(num_list) + 1)
        c.num_list = num_list

        title_list = [f"{i_val[0]}| " + i_val[1] for i_val in c.val_list]
        c.title_list = ["+ Добавить новый скрипт"] + title_list + ["+ Добавить новый скрипт"]
        
        c.choice_variable_num.set(1) 
        c.choice_variable_val.set(c.title_list[0]) 
        c.choice_num_menu.configure(values = c.num_list)
        c.choice_val_menu.configure(values = c.title_list)
        
        c.title_empty.delete(0, END)
        c.text_empty.delete("1.0", END)

        c.btn_save.configure(bg = color["Сохранить"], text = "Сохранить")
        c.btn_delete.configure(state= "disable")

        if status == "Удалить":         
            c.title_empty.insert(0, "Удалено")

            c.title_empty.configure(bg = color["Удалить"])
            c.text_empty.configure(bg = color["Удалить"])

        elif status == "Сохранить":
            c.title_empty.configure(bg = color["Сохранить"])
            c.text_empty.configure(bg = color["Сохранить"])

            c.title_empty.insert(0, "Сохранено")

        elif status == "Изменить":
            c.title_empty.configure(bg = color["Изменить"])
            c.text_empty.configure(bg = color["Изменить"]) 

            c.title_empty.insert(0, "Изменено")

    def switching_assembling_combobox(event):
        """
        Изменяет значения полей title_empty, text_empty, choice_num_menu, btn_save, btn_delete при переключении комбобокса choice_val_menu
        """
        c = Settings_Assembling
        choice = c.choice_val_menu.get()

        if choice == "+ Добавить новый скрипт":
            c.title_empty.delete(0, END)
            c.title_empty.insert(0, "Титульник")

            c.text_empty.delete("1.0", END)
            c.text_empty.insert("1.0", "Текст скрипта\n\n{} - Символ для автоматической вставки")

            c.choice_num_menu.configure(values = c.num_list)

            c.btn_save.configure(text = "Сохранить", bg = color["Сохранить"])
            c.btn_delete.configure(state='disabled')
        else:
            index_title = int(choice[:choice.index('|')])
            
            text = c.val_list[index_title - 1][2]
            c.choice_variable_num.set(index_title) 

            c.title_empty.delete(0, END)
            c.title_empty.insert(0, choice[choice.index(" ") + 1:])

            c.text_empty.delete("1.0", END)
            c.text_empty.insert("1.0", text)
            
            num_list = c.num_list[:len(c.num_list) - 1]
            c.choice_num_menu.configure(values = num_list)

            c.btn_save.configure(text = "Изменить", state='normal', bg = color["Изменить"])
            c.btn_delete.configure(state='normal')
        
        c.title_empty.configure(bg = color["Текст"])
        c.text_empty.configure(bg = color["Текст"]) 



    def assembling_delete():
        """
        Удаляет выбранный скрипт

        choice - Значение choice_val_menu
        index_title - Порядковый номер титульника (14| Титульник) index_title = 14
        num_list - Новый список нумерации скриптов
        title_list - Новый список титульников скриптов

        SQL 1 - Удаляет скрипт с id = index_title
        SQL 2 - Перемещяет все скрипты выше index_title на 1 id ниже
        SQL 3 - Новый список со скриптами
        """
        c = Settings_Assembling
        choice = c.choice_val_menu.get()
        index_title = int(choice[:choice.index('|')])
        title = choice[choice.index(' ') + 1:]
        select_script = cur.execute(f"SELECT id, title FROM assembling_list where id = '{index_title}';").fetchall()[0]

        if select_script == (index_title, title):
            cur.execute(f"DELETE FROM assembling_list where id = '{index_title}';") # SQL 1
            cur.execute(f"UPDATE assembling_list SET id = id - 1 WHERE id > '{index_title}'") # SQL 2
            c.commit_script("Удалить")
        else:
            error_two_open()
    

    def assembling_save():
        """
        Добавляет или изменяет выбранный скрипт

        choice - Значение choice_val_menu
        index_title - Порядковый номер титульника (14| Титульник) index_title = 14
        title - Значение поля title_empty
        text - Значение поля text_empty
        num_list - Новый список нумерации скриптов
        title_list - Новый список титульников скриптов

        SQL 1 - Увеличивает значение id у скрипта где id = i_index
        SQL 2 - Добавляет скрипт в БД
        SQL 3 - Изменяет текст и титульник скрипта с id = index_title
        SQL 4 - Удаляет скрипта с id = index_title
        SQL 5 - Изменяет id скриптов у которых id больше index_title и меньше index_title + 1
        SQL 6 - Новый список скриптов

        Цикл 1 - Добавляет 1 к каждому скрипту чей id = i_index
        """
        c = Settings_Assembling
        choice = choice = c.choice_val_menu.get()
        index = int(c.choice_variable_num.get())
        title = c.title_empty.get()
        text = c.text_empty.get("1.0", END)

        if choice == "+ Добавить новый скрипт":
            max_index = len(c.val_list) + 1

            for i_index in range(max_index, index - 1, -1):  # Цикл 1     
                cur.execute(f"UPDATE assembling_list SET id = id + 1 WHERE id = '{i_index}'") # SQL 1
                conn.commit()

            cur.execute("INSERT INTO assembling_list(id, title, text) VALUES (?, ?, ?)", (index, title, text[:-1])) # SQL 2
            conn.commit()

            c.commit_script("Сохранить")

        else:
            index_title = int(choice[:choice.index('|')])

            title = choice[choice.index(' ') + 1:]
            select_script = cur.execute(f"SELECT id, title FROM assembling_list where id = '{index_title}';").fetchall()[0]
            
            if select_script == (index_title, title):
                if index_title == index:
                    cur.execute(f"UPDATE assembling_list SET title = ?, text = ? WHERE id = ?", (title, text[:-1], index_title)) # SQL 3
                    
                    
                elif index_title < index:    
                    cur.execute(f"DELETE FROM assembling_list where id = '{index_title}';") # SQL 4
                    cur.execute(f"UPDATE assembling_list SET id = id - 1 WHERE id > '{index_title}' AND id <= '{index}'") # SQL 5
                    cur.execute("INSERT INTO assembling_list(id, title, text) VALUES (?, ?, ?)", (index, title, text[:-1])) # SQL 2
                    

                else:
                    cur.execute(f"DELETE FROM assembling_list where id = '{index_title}';") # SQL 4
                    
                    for i_index in range(index_title - 1, index - 1, -1): # Цикл 2  
                        cur.execute(f"UPDATE assembling_list SET id = id + 1 WHERE id = '{i_index}'") # SQL 1
                        conn.commit()

                    cur.execute("INSERT INTO assembling_list(id, title, text) VALUES (?, ?, ?)", (index, title, text[:-1])) # SQL 2
                    
                c.commit_script("Изменить")
            else:
                error_two_open()

    """ 
    /////// Фреймы /////// 

    frame_choice - Фрейм для выбора названия скрипта
    frame_title - Фрейм для имени скрипта
    frame_text - Фрейм для текста скрипта
    frame_btn - Фрейм для кнопок удалить сохранить изменить

    frame_choice_num - Дочерний фрейм frame_choice для нумерации скрипта
    frame_choice_val - Дочерний фрейм frame_choice для названий скриптов

    /////// Элементы /////// 

    choice_variable_num - Значение в комбобоксе choice_num_menu
    choice_num_menu - комбобокс номеров скриптов
    choice_variable_val - Значение в комбобоксе choice_val_menu
    choice_val_menu - комбобокс значений скриптов
    title_empty - Поле ввода титульника
    text_empty - поле ввода текста скрипта
    btn_delete - кнопка удаления скрипта
    btn_save - кнопка изменения или добавления скрипта

    /////// Переменные /////// 

    val_list - список скриптов
    num_list - Список номеров скриптов
    title_list - Список титульников скриптов
    """

    frame_choice = Frame(Settings.frame_assembling)
    frame_choice.pack(fill=X, pady = 5)

    frame_title = Frame(Settings.frame_assembling)
    frame_title.pack(fill=X, pady = 5)

    frame_text = Frame(Settings.frame_assembling)
    frame_text.pack(fill=X, pady = 5)

    frame_btn = Frame(Settings.frame_assembling)
    frame_btn.pack(fill=BOTH)

    frame_choice_num = Frame(frame_choice)
    frame_choice_num.pack(anchor=NW, side=LEFT, fill=X)

    frame_choice_val = Frame(frame_choice)
    frame_choice_val.pack(anchor=NW, side=LEFT, fill=X)

    cur.execute("SELECT id, title, text FROM assembling_list;")
    val_list = cur.fetchall()

    num_list = [i_val[0] for i_val in val_list]
    num_list.append(len(num_list) + 1)
    title_list = [f"{i_val[0]}| " + i_val[1] for i_val in val_list]
    title_list = ["+ Добавить новый скрипт"] + title_list + ["+ Добавить новый скрипт"]

    choice_variable_num = StringVar(frame_choice_num)
    choice_variable_num.set(1) 
    choice_num_menu = ttk.Combobox(frame_choice_num, width=3, state = "readonly", font = font["Средний"], textvariable=choice_variable_num, values=num_list)
    choice_num_menu.pack()

    choice_variable_val = StringVar(frame_choice_val)
    choice_variable_val.set(title_list[0]) 
    choice_val_menu = ttk.Combobox(frame_choice_val, width=100, state = "readonly", font = font["Средний"], textvariable=choice_variable_val, values=title_list)
    choice_val_menu.pack()
    choice_val_menu.bind("<<ComboboxSelected>>", switching_assembling_combobox)

    title_lb = Label(frame_title, text = "Название")
    title_lb.pack(fill=X)
    title_empty = Entry(frame_title, font = font["Большой"])
    title_empty.pack(fill=X)
    title_empty.bind(ctrl, keys)

    text_lb = Label(frame_text, text = "Текст")
    text_lb.pack(fill=X)
    text_empty = Text(frame_text, wrap=WORD, undo=True, font = font["Средний"])
    text_empty.pack(side=LEFT, expand=True, fill = BOTH)
    text_empty.bind(ctrl, keys)

    btn_delete = Button(frame_btn,text = "Удалить", state= "disable", bg = color["Удалить"], command = assembling_delete)
    btn_delete.pack(side=LEFT, expand=True, fill="both")
    
    btn_save = Button(frame_btn, text = "Сохранить", bg = color["Сохранить"], command = assembling_save)
    btn_save.pack(side=LEFT, expand=True, fill="both")


class Settings_Ready:

    def commit_script(status):
        c = Settings_Ready
        conn.commit()
        cur.execute(f"SELECT num, title FROM ready_category ORDER BY num;")
        val_list = cur.fetchall()

        category_num_list = [i_val[0] for i_val in val_list]
        category_num_list.append(len(category_num_list) + 1)
        c.category_num_list = category_num_list

        category_list = [f"{i_val[0]}| " + i_val[1] for i_val in val_list]
        category_list = ["+ Добавить новую категорию"] + category_list + ["+ Добавить новую категорию"]
        c.category_list = category_list

        c.choice_category_num_menu.configure(values=c.category_num_list)
        c.choice_category_menu.configure(values=c.category_list)

        c.choice_category_num.set("1") 
        c.choice_category_val.set(category_list[0])

        c.choice_script_num.set("") 
        c.choice_script_val.set("")

        c.choice_script_menu.configure(state='disabled')
        c.choice_script_num_menu.configure(state='disabled')
        
        c.title_empty.delete(0, END)
        c.text_empty.delete("1.0", END)

        c.btn_save.configure(bg = color["Сохранить"], text = "Сохранить")
        c.btn_delete.configure(state="disabled")

        if status == "Удалить":         
            c.title_empty.insert(0, "Удалено")

            c.title_empty.configure(bg = color["Удалить"])
            c.text_empty.configure(bg = color["Удалить"])

        elif status == "Сохранить":
            c.title_empty.configure(bg = color["Сохранить"])
            c.text_empty.configure(bg = color["Сохранить"])

            c.title_empty.insert(0, "Сохранено")

        elif status == "Изменить":
            c.title_empty.configure(bg = color["Изменить"])
            c.text_empty.configure(bg = color["Изменить"]) 

            c.title_empty.insert(0, "Изменено")

        elif status == "Ошибка создания":
            c.title_empty.insert(0, "Имя занято")
            c.btn_save.configure(state="disabled")
            c.title_empty.configure(bg = color["Удалить"])
            c.text_empty.configure(bg = color["Удалить"])


    def new_category_btn():
        c = Ready
        cur.execute("""SELECT title FROM ready_category ORDER BY num;""")
        ready_category = cur.fetchall()

        for i_val in c.category_list:
            i_val.destroy()

        c.category_list.clear()
        c.with_list = 0
        for i_val in ready_category: # Цикл 4
        
            val = i_val[0]
            
            width_btn = int(len(str(val)))
            c.with_list += width_btn

            btn_readys = Button(c.Сanvas_frame, text = val, width = width_btn,  bg = color["Активная"], command = lambda val=val: Ready.ready_swap_category(val)) 
            btn_readys.pack(side=LEFT, expand=True, fill="both")
            c.category_list.append(btn_readys)

        c.category_list[0].configure(state='disabled', bg = color["Активная категория"])
        btn_readys_end = Button(c.Сanvas_frame, text = "             ", state='disabled', bg = color["Активная"]) 
        btn_readys_end.pack(side=LEFT, expand=True, fill="both")
        c.category_list.append(btn_readys_end)
        c.Canvas_top.config(scrollregion=(0, 0, c.with_list*12, 0))

    def switching_category_combobox(event):
        c = Settings_Ready
        choice = c.choice_category_val.get()

        if choice == "+ Добавить новую категорию":
            c.title_empty.delete(0, END)
            c.title_empty.insert(0, "Титульник")

            c.choice_script_num_menu.configure(state='disabled')
            c.choice_script_menu.configure(state='disabled')

            c.choice_category_num_menu.configure(values = c.category_num_list)
            
            c.btn_save.configure(text = "Сохранить", bg = color["Сохранить"], state="normal")
            c.btn_delete.configure(state='disabled')
        else:
            try:
                index_title = int(choice[:choice.index('|')])
                cur.execute(f"SELECT id FROM ready_category WHERE num = '{index_title}';")
                id_category = cur.fetchone()[0]

                cur.execute(f"SELECT id, title, text FROM ready_text WHERE id_category = {id_category} ORDER BY id;")
                c.script_list = cur.fetchall()

                c.script_num_list = [i_val[0] for i_val in c.script_list]
                c.script_num_list.append(len(c.script_num_list) + 1)
                c.script_title_list = [f"{i_val[0]}| " + i_val[1] for i_val in c.script_list]
                c.script_title_list = ["+ Добавить новый скрипт", ""] + c.script_title_list + ["", "+ Добавить новый скрипт"]
                
                category_num_list = c.category_num_list[:len(c.category_num_list)-1]
                c.choice_category_num_menu.configure(values = category_num_list)

                c.choice_script_num_menu.configure(values = c.script_num_list)
                c.choice_script_menu.configure(values = c.script_title_list , state='readonly')

                c.choice_category_num.set(index_title) 
        
                c.title_empty.delete(0, END)
                c.title_empty.insert(0, choice[choice.index(" ") + 1:])

                c.btn_save.configure(text = "Изменить", state='normal', bg = color["Изменить"])
                c.btn_delete.configure(state='normal')
            except TypeError:
                error_two_open()
        
        c.choice_script_num.set("") 
        c.choice_script_val.set("") 
        c.choice_script_num_menu.configure(state='disabled')
        c.text_empty.delete("1.0", END)
        c.text_empty.configure(state="disabled")
        
        c.title_empty.configure(bg = color["Текст"])
        c.text_empty.configure(bg = color["Текст"])


    def ready_delete():
        c = Settings_Ready

        choice_category = c.choice_category_val.get()
        title_category = choice_category[choice_category.index(' ') + 1:]
        choice_script = c.choice_script_val.get()
        index_category = int(choice_category[:choice_category.index('|')])
        select_category = cur.execute(f"SELECT title FROM ready_category WHERE num = '{index_category}';").fetchall()
        try:
            if choice_script == "" and select_category[0][0] == title_category:
                    conn.execute("PRAGMA foreign_keys = 1")
                    cur.execute(f"DELETE FROM ready_category WHERE num = '{index_category}';")
                    cur.execute(f"UPDATE ready_category SET num = num - 1 WHERE num > '{index_category}'")
                    conn.commit()
                    c.new_category_btn()
                    c.commit_script("Удалить")   


            elif choice_script != "":
                index_script = int(choice_script[:choice_script.index('|')])    
                id_category =  cur.execute(f"SELECT id FROM ready_category WHERE num = '{index_category}'").fetchone()[0]

                select_script = cur.execute(f"SELECT id, title FROM ready_text WHERE id_category = '{id_category}' AND id = {index_script};").fetchall()[0]
                num_script = choice_script[choice_script.index(' ') + 1:]

                if select_script == (index_script, num_script) and select_category[0][0] == title_category:
                    cur.execute(f"DELETE FROM ready_text WHERE id_category = '{id_category}' AND id = '{index_script}';")
                    cur.execute(f"UPDATE ready_text SET id = id - 1 WHERE id_category = '{id_category}' AND id > '{index_script}'")
                    conn.commit()    
                    c.commit_script("Удалить")   
                else:
                    error_two_open()

            else:
                error_two_open()
        except IndexError:
            error_two_open()


    def ready_save():
        c = Settings_Ready
        choice_category = c.choice_category_val.get()
        choice_script = c.choice_script_val.get()
        title = c.title_empty.get()
        text = c.text_empty.get("1.0", END)
        category_num = int(c.choice_category_num.get())
        script_num = c.choice_script_num.get()
        
        cur.execute("SELECT title FROM ready_category WHERE title = ?;", (title,))
        category_choise = cur.fetchall()

        if choice_category == "+ Добавить новую категорию":
            if len(category_choise) == 0:
                cur.execute(f"UPDATE ready_category SET num = num + 1 WHERE num >= '{category_num}'")
                cur.execute("INSERT INTO ready_category(num, title) VALUES (?, ?)", (category_num, title))
                
                c.commit_script("Сохранить")
                c.new_category_btn()
            else:
                c.commit_script("Ошибка создания")
        
        elif choice_script == "":
            index_category = int(choice_category[:choice_category.index('|')])
            title_val = choice_category[choice_category.index(" ") + 1:]
            select_category = cur.execute(f"SELECT num, title FROM ready_category WHERE num = '{index_category}';").fetchall()

            if (title_val == title or len(category_choise)) == 0 and select_category[0] == (index_category, choice_category[choice_category.index(' ') + 1:]):

                if index_category == category_num:
                    cur.execute(f"UPDATE ready_category SET title = ? WHERE num = ?", (title, category_num))
                    
                elif index_category > category_num:
                    cur.execute(f"UPDATE ready_category SET num = num + 1 WHERE num >= '{category_num}' AND num < '{index_category}'")
                    cur.execute(f"UPDATE ready_category SET num = ?, title = ? WHERE title = ?", (category_num, title, title_val))
                    
                elif index_category < category_num:
                    cur.execute(f"UPDATE ready_category SET num = num - 1 WHERE num <= '{category_num}' AND num > '{index_category}'")
                    cur.execute(f"UPDATE ready_category SET num = ?, title = ? WHERE title = ?", (category_num, title, title_val))  
                
                c.commit_script("Изменить")
                c.new_category_btn()
            elif select_category[0] != (index_category, choice_category[choice_category.index(' ') + 1:]):
                error_two_open()
            else:
                c.commit_script("Ошибка создания")
            

        elif choice_script == "+ Добавить новый скрипт":
                title_category = choice_category[choice_category.index(' ') + 1:]
                index_script = int(choice_script[:choice_script.index('|')])    
                id_category =  cur.execute(f"SELECT id FROM ready_category WHERE num = '{index_category}'").fetchone()[0]

                select_script = cur.execute(f"SELECT id, title FROM ready_text WHERE id_category = '{id_category}' AND id = {index_script};").fetchall()[0]
                num_script = choice_script[choice_script.index(' ') + 1:]

                if select_script == (index_script, num_script) and select_category[0][0] == title_category:
                    index_category = int(choice_category[:choice_category.index('|')])
                    cur.execute(f"SELECT id FROM ready_category WHERE num = '{index_category}'")
                    id_category = cur.fetchone()[0]

                    cur.execute(f"UPDATE ready_text SET id = id + 1 WHERE id >= '{script_num}' AND id_category = {id_category}")
                    cur.execute("INSERT INTO ready_text(id, id_category, title, text) VALUES (?, ?, ?, ?)", (script_num, id_category, title, text[:-1]))
                    conn.commit()

                    c.commit_script("Сохранить")
                else:
                    error_two_open()

        elif choice_script != "+ Добавить новый скрипт":
            index_category = int(choice_category[:choice_category.index('|')])
            index_script = int(choice_script[:choice_script.index('|')])
            
            script_num = int(script_num)

            cur.execute(f"SELECT id FROM ready_category WHERE num = '{index_category}'")
            id_category = cur.fetchone()[0]

            title_category = choice_category[choice_category.index(' ') + 1:]
            id_category =  cur.execute(f"SELECT id FROM ready_category WHERE num = '{index_category}'").fetchone()[0]

            select_script = cur.execute(f"SELECT id, title FROM ready_text WHERE id_category = '{id_category}' AND id = {index_script};").fetchall()[0]
            num_script = choice_script[choice_script.index(' ') + 1:]

            if select_script == (index_script, num_script):
                if script_num == index_script:
                    cur.execute(f"UPDATE ready_text SET title = ?, text = ? WHERE id = ? AND id_category = {id_category}", (title, text[:-1], script_num))
                    
                elif script_num > index_script:
                    cur.execute(f"DELETE FROM ready_text WHERE id = '{index_script}' AND id_category = {id_category};")
                    cur.execute(f"UPDATE ready_text SET id = id - 1 WHERE id > '{index_script}' AND id <= '{script_num}' AND id_category = {id_category}")
                    cur.execute("INSERT INTO ready_text(id, id_category, title, text) VALUES (?, ?, ?, ?)", (script_num, id_category, title, text[:-1]))
                
                elif script_num < index_script:
                    cur.execute(f"DELETE FROM ready_text WHERE id = '{index_script}' AND id_category = {id_category};")
                    cur.execute(f"UPDATE ready_text SET id = id + 1 WHERE id < '{index_script}' AND id >= '{script_num}' AND id_category = {id_category}")
                    cur.execute("INSERT INTO ready_text(id, id_category, title, text) VALUES (?, ?, ?, ?)", (script_num, id_category, title, text[:-1]))
                
                c.commit_script("Изменить")
            else:
                error_two_open()
        else:
            error_two_open()



    def switching_script_combobox(event):
        c = Settings_Ready 
        choice_script = c.choice_script_val.get()
        choice_category = c.choice_category_val.get()

        c.title_empty.delete(0, END)
        c.text_empty.delete("1.0", END)

        if choice_script == "":
            c.choice_script_num_menu.configure(state='disabled')
            c.text_empty.configure(state="disabled")
            c.choice_script_num.set("")

            c.title_empty.insert(0, choice_category[choice_category.index(" ") + 1:])
        
            c.btn_delete.configure(state="normal")
            c.btn_save.configure(text="Изменить", bg=color["Изменить"])

        elif choice_script == "+ Добавить новый скрипт":
            c.choice_script_num_menu.configure(state="readonly")
            c.choice_script_num_menu.configure(state="readonly", values = c.script_num_list)

            c.text_empty.configure(state="normal")
            if c.choice_script_num.get() == "":
                c.choice_script_num.set(1)

            c.title_empty.insert(0, "Титульник")
            c.text_empty.insert("1.0", "Текст скрипта\n\n{} - Символ для автоматической вставки")


            c.btn_delete.configure(state="disabled")
            c.btn_save.configure(text="Сохранить", bg=color["Сохранить"])
        else:
            script_num = int(choice_script[:choice_script.index('|')])

            script_num_list = c.script_num_list[:len(c.script_num_list)-1]
            c.choice_script_num_menu.configure(state="readonly", values = script_num_list)

            c.text_empty.configure(state="normal")
            c.choice_script_num.set(script_num)

            c.title_empty.insert(0, choice_script[choice_script.index(" ") + 1:])
            c.text_empty.insert("1.0", c.script_list[script_num - 1][2])

            c.btn_delete.configure(state="normal")
            c.btn_save.configure(text="Изменить", bg=color["Изменить"])

        c.title_empty.configure(bg = color["Текст"])
        c.text_empty.configure(bg = color["Текст"])
    """ 
    /////// Фреймы /////// 

    /////// Элементы /////// 

    /////// Переменные /////// 

    """
    
    frame_choice_category = Frame(Settings.frame_Ready)
    frame_choice_category.pack(fill=X)
        
    frame_choice_script = Frame(Settings.frame_Ready)
    frame_choice_script.pack(fill=X, pady = 5)

    frame_title = Frame(Settings.frame_Ready)
    frame_title.pack(fill=X, pady = 5)

    frame_text = Frame(Settings.frame_Ready)
    frame_text.pack(fill=X, pady = 5)

    frame_btn = Frame(Settings.frame_Ready)
    frame_btn.pack(fill=BOTH)

    frame_choice_category_num = Frame(frame_choice_category)
    frame_choice_category_num.pack(anchor=NW, side=LEFT, fill = X)

    frame_choice_category_val = Frame(frame_choice_category)
    frame_choice_category_val.pack(anchor=NW, side=LEFT, fill = X)
    
    frame_choice_script_num = Frame(frame_choice_script)
    frame_choice_script_num.pack(anchor=NW, side=LEFT, fill = X)

    frame_choice_script_val = Frame(frame_choice_script)
    frame_choice_script_val.pack(anchor=NW, side=LEFT, fill = X)

    """--- Раздел с категориями  ---"""

    cur.execute("SELECT num, title FROM ready_category ORDER BY num;")
    category_list = cur.fetchall()

    category_num_list = [i_val[0] for i_val in category_list]
    category_num_list.append(len(category_num_list) + 1)
    category_title_list = [f"{i_val[0]}| " + i_val[1] for i_val in category_list]
    category_title_list = ["+ Добавить новую категорию"] + category_title_list + ["+ Добавить новую категорию"]
    category_id = 1
    
    choice_category_num = StringVar(frame_choice_category_num)
    choice_category_num.set(1) 
    choice_category_num_menu = ttk.Combobox(frame_choice_category_num, width=3, state = "readonly", font = font["Средний"], textvariable=choice_category_num, values=category_num_list)
    choice_category_num_menu.pack()

    choice_category_val = StringVar(frame_choice_category_val)
    choice_category_val.set(category_title_list[0]) 
    choice_category_menu = ttk.Combobox(frame_choice_category_val, width=100, state = "readonly", font = font["Средний"], textvariable=choice_category_val, values=category_title_list)
    choice_category_menu.pack()
    choice_category_menu.bind("<<ComboboxSelected>>", switching_category_combobox)
    
    """--- Раздел с скриптами ---"""
    script_num_list = list()
    script_title_list = list()

    script_list = list()

    choice_script_num = StringVar(frame_choice_script_num)
    choice_script_num.set("") 
    choice_script_num_menu = ttk.Combobox(frame_choice_script_num, width=3, state = "disabled", textvariable=choice_script_num)
    choice_script_num_menu.pack()

    choice_script_val = StringVar(frame_choice_script_val)
    choice_script_val.set("") 
    choice_script_menu = ttk.Combobox(frame_choice_script_val, width=100, state = "disabled", textvariable=choice_script_val)
    choice_script_menu.pack()
    choice_script_menu.bind("<<ComboboxSelected>>", switching_script_combobox)

    """--- Раздел полями ввода  ---"""
    title_lb = Label(frame_title, text = "Название")
    title_lb.pack(fill=X)
    title_empty = Entry(frame_title, width=100, bg = color["Текст"], font = font["Большой"])
    title_empty.pack()
    title_empty.bind(ctrl, keys)

    text_lb = Label(frame_text, text = "Текст")
    text_lb.pack(fill=X)
    text_empty = Text(frame_text, state= "disable", bg = color["Текст"], font = font["Средний"], wrap=WORD, undo=True)
    text_empty.pack()
    text_empty.bind(ctrl, keys)

    btn_delete = Button(frame_btn, text = "Удалить", state= "disable", bg = color["Удалить"], command=ready_delete)
    btn_delete.pack(side=LEFT, expand=True, fill="both")
    
    btn_save = Button(frame_btn, text = "Сохранить", bg = color["Сохранить"], command=ready_save)
    btn_save.pack(side=LEFT, expand=True, fill="both")


class Book:
    """
    Вкладка Блокнота
    book_text - Поле для блокнота растянутое на всю ширину вкладки
    """
    def save_to_database(event=None):
        c = Book
        new_note = c.book_important_text.get("1.0", "end-1c")

        cur.execute("UPDATE setting SET important_book_text = ? WHERE id = 1", (new_note,))
        conn.commit()

    cur.execute("SELECT important_book, important_book_text FROM setting;")
    setting = cur.fetchall()

    book_text = Text(main.frame_book, wrap=WORD, undo=True, font = font["Большой"])
    book_text.pack(expand=True, fill = BOTH)
    book_text.bind(ctrl, keys)

    book_important_text = Text(main.frame_book, height=10, font = font["Большой"], state= "normal", wrap=WORD, undo=True)
    book_important_text.insert("1.0", setting[0][1])
    book_important_text.bind(ctrl, keys)
    book_important_text.bind('<KeyRelease>', save_to_database)

    if setting[0][0] == 1:
        book_important_text.pack(fill = X)
    else:
        book_important_text.pack_forget()
    

class Settings_general:

    def export_file():
        category = cur.execute("""SELECT id, title FROM ready_category ORDER BY num""").fetchall()
        assembling = cur.execute("""SELECT title, text FROM assembling_list""").fetchall()

        exel = Workbook()
        my_exel = exel.active
        list_exel = exel.sheetnames

        exel.create_sheet("Сборочные")
        exel_assembling = exel["Сборочные"]

        for i_num, i_val in enumerate(assembling):
            title = exel_assembling.cell(row=i_num + 1, column=1)
            title.value = i_val[0]

            text = exel_assembling.cell(row=i_num + 1, column=2)
            text.value = i_val[1]

        for i_val in category:
            category_id = i_val[0]
            category_title = i_val[1]


            cur.execute(f"""SELECT title, text FROM ready_text WHERE id_category = {category_id} ORDER BY id;""")
            script = cur.fetchall()

            exel.create_sheet(str(category_title))
            exel_category = exel[category_title]

            for i_num, i_val in enumerate(script):

                title = exel_category.cell(row=i_num + 1, column=1)
                title.value = i_val[0]

                text = exel_category.cell(row=i_num + 1, column=2)
                text.value = i_val[1]

        exel.remove(exel[list_exel[0]])
        exel.save('my_script.xlsx')

    def book_checkbutton_changed():
        
        cur.execute("SELECT important_book FROM setting WHERE id = 1;")
        important_book = cur.fetchone()

        if important_book[0] == 1:
            Book.book_important_text.pack_forget()
            new = 0
        else:
            Book.book_important_text.pack(fill = X)
            new = 1

        cur.execute(f"UPDATE setting SET important_book = ? WHERE id = 1;", (new,))
        conn.commit()

    if Book.setting[0][0] == 1:
        num = 1
    else: 
        num = 0

    var = IntVar(value=num)
    book_checkbutton = Checkbutton(Settings.frame_general, text="Блокнот для заметок", font = font["Большой"], variable=var, onvalue=1, offvalue=0, command=book_checkbutton_changed)
    book_checkbutton.pack(padx=6, pady=6, anchor=NW)

    link_frame = Frame(Settings.frame_general)
    link_frame.pack(fill=X, side=BOTTOM)

    entry_link = Text(link_frame, height=2, state="disable")
    entry_link.pack(fill=X)

    link_frame_btn = Frame(link_frame)
    link_frame_btn.pack(fill=X)

    export_btn = Button(link_frame_btn, text = "Выгрузить", font = font["Большой"], command=export_file)
    export_btn.pack(fill=X, expand=True, side=LEFT)
    
    import_btn = Button(link_frame_btn, text = "Загрузить", font = font["Большой"])
    import_btn.pack(fill=X, expand=True, side=LEFT)

root.mainloop()
                                        