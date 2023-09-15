from tkinter import *
from tkinter import ttk
from openpyxl import Workbook
import pyperclip
import sqlite3
import ctypes
import math
import sys
import os

color = {
    "Активная":"#F7C815",
    "Неактивная":"#deb416",
    "Сохранить":"#228b22",
    "Изменить":"#FAD074",
    "Удалить":"#ff2400",
    "Активная категория":"#EC9704",
    "Текст":"#fff",
    "Обводка_сборочных":"#F7E8E3",
}

font = {
    "Большой":('Helvetica', 12),
    "Средний":('Helvetica', 10),
}

my_file = os.path.abspath(sys.argv[0]) 
my_dir = os.path.dirname(my_file)
db_file = os.path.join(my_dir, "orders.db") # Ссылка БД
conn = sqlite3.connect(db_file)
cur = conn.cursor()

if sys.platform == 'win32' or sys.platform.startswith('linux'):
    ctrl = "<Control-KeyPress>"
elif sys.platform == "darwin":
    ctrl = "<Command-KeyPress>"

def get_path(event):
    print(event.data)

def error_two_open():
    error = Tk()
    error.title("TCRM+")
    error.geometry("400x200")
    error.attributes("-topmost",True)
    error.iconbitmap("icon.ico")

    error_label = Label(error, text = "ОШИБКА:\nНе удалось отправить запрос на базу данных.\nЗакройте все окошки приложения и откройте только одно.\nТогда ошибка не повторится")
    error_label.pack()
    root.destroy()

def is_ru_lang_keyboard():
    """
    Функция для работы функции keys. Для работы блокнота в русской раскладке
    """
    if sys.platform == 'win32':
        u = ctypes.windll.LoadLibrary("user32.dll")
        pf = getattr(u, "GetKeyboardLayout")
        return hex(pf(0)) == '0x4190419'
    elif sys.platform == "darwin":
        u = ctypes.cdll.LoadLibrary("/System/Library/Frameworks/ApplicationServices.framework/Versions/A/Frameworks/HIServices.framework/Versions/A/HIServices")
        pf = getattr(u, "TISGetInputSourceProperty")
        property_name = ctypes.c_void_p.in_dll(u, "kTISPropertyUnicodeKeyLayoutData")
        handle = pf(ctypes.c_void_p.in_dll(u, "kTISPropertyInputSourceClass"), property_name)
        return bool(handle)
    elif sys.platform.startswith('linux'):
        u = ctypes.cdll.LoadLibrary('libX11.so')
        display = u.XOpenDisplay(None)
        if display == 0:
            return False
        keysym = u.XStringToKeysym('F2')
        keycode = u.XKeysymToKeycode(display, keysym)
        return bool(keycode)
    return False 

def keys(event):
    if is_ru_lang_keyboard():
        if event.keycode==86: 
            event.widget.event_generate("<<Paste>>")

        elif event.keycode==67: 
            pyperclip.copy(event.widget.selection_get())

        elif event.keycode==88:  
            pyperclip.copy(event.widget.selection_get())
            event.widget.delete("sel.first", "sel.last")
            
        elif event.keycode==65535: 
            event.widget.event_generate("<<Clear>>")

        elif event.keycode==65: 
            event.widget.event_generate("<<SelectAll>>")

        elif event.keycode == 90:
            event.widget.edit_undo()
        
        elif event.keycode == 8:
            event.widget.delete("insert-1c wordstart", "insert")
            
    elif event.keycode==67: 
        pyperclip.copy(event.widget.selection_get())
        return "break"
    
    elif event.keycode==88:  
        pyperclip.copy(event.widget.selection_get())
        event.widget.delete("sel.first", "sel.last")
        return "break"
    
    elif event.keycode == 8:
        event.widget.delete("insert-1c wordstart", "insert")
        

def copy(val, buffer = None):
    """
    Копирование данных в буфер обмена  
    """

    pyperclip.copy(val.format(main.aut_text_entry.get())) #  копирует val в буфер обмена. val - Копируемый текст

    if buffer != None: # buffer - Текстовое поле где отражен скопированный текст
        buffer.delete("1.0", END)
        buffer.insert("2.0", val.format(main.aut_text_entry.get()))

root = Tk() # Основная рамка приложения
root.title("TCRM+") # Название приложения
root.geometry("265x673") # Размеры приложения
root.minsize(width=200, height=640) # Минимальные размеры приложения
root.maxsize(width=300, height=700) # Максимальный раззмеры приложения
root.attributes("-topmost",True) # Поверх всех окон


icon_file = os.path.join(my_dir, "icon.ico") # Ссылка на иконку приложения
if os.path.exists(icon_file): # Если сылка найдена то отобразить
    root.iconbitmap(icon_file)

class main:

    """
    Основная панель, содержит: 
    Поле вставки из буфера, 
    счетчик,
    Вкладки.

    /////// Функции ///////
    past - Вставить из буфера в текстовое поле

    f_counter_minus, f_counter_plus, f_counter_null - Функции контроля счетчика +, -, 0
    """

    def past():

        """
        Вставляет в текстовое поле содержимое буфера обмена
        """

        c = main
        empty_str = pyperclip.paste() # Содержание буфера обмена
        empty_new = str() # Новый текст если в конце empty_str есть пробел
        if empty_str[-1] == " " and len(empty_str) > 0:
            empty_new = empty_str[0:len(empty_str)-1]
        else:
            empty_new = empty_str
        
        c.aut_text_entry.delete(0, END) # Удалить все и вставить empty_new в тектовое поле 
        c.aut_text_entry.insert(0, empty_new)


    def f_counter_minus():
        
        """
        Убавляет значение 1 со счетчика
        """
        c = main

        if c.counter["text"] > "0":

            cur.execute(""" 
            UPDATE counter 
            SET count = count - 1""") # Уменьшает счетчик в БД
            conn.commit()

            new_empty = str(int(c.counter["text"]) - 1) # Новое значение счетчика
            c.counter["text"] = new_empty


    def f_counter_plus():

        """
        Добавляет к счетчику значение 1
        """
        c = main
        cur.execute("""
        UPDATE counter 
        SET count = count + 1""") # Увеличивает счетчик в БД
        conn.commit()

        new_empty = str(int(c.counter["text"]) + 1) # Новое значение счетчика
        c.counter["text"] = new_empty


    def f_counter_null():
        
        """
        Обнуляет счетчик
        """
        c = main
        cur.execute("""
        UPDATE counter 
        SET count = 0""") # Обнуляет счетчик в БД
        conn.commit()

        c.counter["text"] = "0"


    frame = Frame(root) # Основной фрейм растянутый на все поле
    frame.place(rely=0, relheight=1, relwidth=1)

    frame_aut_text = Frame(frame) # Фрейм для вставки
    frame_aut_text.pack(fill=X)

    frame_counter = Frame(frame) # Фрейм счетчика
    frame_counter.pack(fill=X)

    frame_bot = Frame(frame) # Фрейм Вкладок - растянут на всю доступную высоту
    frame_bot.pack(fill=X)

    aut_text_btn = Button(frame_aut_text, text = "Вставить", width=10, command= lambda: main.past()) # Кнопка вставки в поле aut_text_entry
    aut_text_btn.grid(column=0, row = 0)

    aut_text_entry = Entry(frame_aut_text, width=100) # Текстовое поле. С него копируют текст для скрипта(обычно имя)
    aut_text_entry.grid(column=1, row = 0)
    aut_text_entry.bind(ctrl, keys)

    counter_plus = Button(frame_counter, text = "+", width=10, command = f_counter_plus) # Кнопки контроля счетчика +1
    counter_plus.grid(column=0, row = 0, sticky="nswe")

    counter_minus = Button(frame_counter, text = "-", width=10, command = f_counter_minus) # Кнопки контроля счетчика -1
    counter_minus.grid(column=1, row = 0, sticky="nswe")

    counter_null = Button(frame_counter, text = "0", width=1, command = f_counter_null) # Кнопки контроля счетчика 0
    counter_null.grid(column=2, row = 0)

    cur.execute("SELECT count FROM counter")
    count = str(cur.fetchone()[0]) # значение счетчика из базы данных
    counter = Label(frame_counter, text= count, font = (font["Большой"], 12)) # Поле ввода счетчика
    counter.grid(column=3, row = 0)

    notebook = ttk.Notebook(frame_bot) # Класс вкладок
    notebook.pack(expand=True, fill=BOTH)

    # Фреймы для вкладок
    frame_assembling = ttk.Frame(notebook)
    frame_Ready = ttk.Frame(notebook)
    frame_settings = ttk.Frame(notebook)
    frame_book = ttk.Frame(notebook)


    frame_assembling.pack(fill=BOTH, expand=True)
    frame_Ready.pack(fill=BOTH, expand=True)
    frame_settings.pack(fill=BOTH, expand=True)
    frame_book.pack(fill=BOTH, expand=True)

    notebook.add(frame_assembling, text="СБОРКА")
    notebook.add(frame_Ready, text="ГОТОВЫЕ")
    notebook.add(frame_settings, text="НАСТРОЙКА")
    notebook.add(frame_book, text="Б")


class Ready:
    """
    Вкладка готовых скриптов. Содержит вкладки категорий скриптов и кнопки для копирования скриптов
    
    /////// Функции ///////
    
    ready_next_page, ready_first_page, ready_back_page - Управление страницами скриптов +1, -1, 1
    ready_swap_category - Переключение категорий при нажатии на кнопку категории
    """


    def ready_next_page():
        """
        Переключение страниц скриптов вперед для вкладки готовых скриптов
        """

        c = Ready

        if c.page != c.len_ready_text:
            c.page += 1
            c.page_list += 10

            c.center_btn.configure(text = f"{c.page}/{c.len_ready_text}")
            
            cur.execute(f"""
                        SELECT ready_text.title, ready_text.text
                        FROM ready_text
                        JOIN ready_category ON ready_category.id = ready_text.id_category
                        WHERE ready_category.title = "{c.category}"
                        ORDER BY ready_text.id
                        LIMIT 10 OFFSET {c.page_list};
                        """)
            val_list = cur.fetchall() # Новый список скриптов

            for i_num, i_val in enumerate(c.btn_script_ready_list): # Переименование текста и закрепленных за кнопкой скриптов
                try:

                    key = val_list[i_num][0] # Титульник скрипта
                    val = val_list[i_num][1] # Текст скрипта
                    
                    i_val.configure(text = key, state='normal', command = lambda val=val: copy(val))

                except IndexError: # Если в списке закончились значения то кнопка неактивна
                    i_val.configure(text = "", state='disabled', bg = color["Неактивная"])


    def ready_back_page():
        """
        Переключение страниц скриптов назад для вкладки готовых скриптов
        """

        c = Ready

        if c.page != 1:
            c.page -= 1
            c.page_list -= 10

            c.center_btn.configure(text = f"{c.page}/{c.len_ready_text}")
            
            cur.execute(f"""
                        SELECT ready_text.title, ready_text.text
                        FROM ready_text
                        JOIN ready_category ON ready_category.id = ready_text.id_category
                        WHERE ready_category.title = "{c.category}"
                        ORDER BY ready_text.id
                        LIMIT 10 OFFSET {c.page_list};
                        """)
            val_list = cur.fetchall() # Новый список скриптов

            for i_num, i_val in enumerate(c.btn_script_ready_list): # Переименование текста и закрепленных за кнопкой скриптов
                try:
                    key = val_list[i_num][0] # Титульник скрипта
                    val = val_list[i_num][1] # Текст скрипта

                    i_val.configure(text = key, bg = color["Активная"], state='normal', command = lambda val=val: copy(val))
                except IndexError:
                    i_val.configure(text = "", state='disabled', bg = color["Неактивная"])


    def ready_first_page():
        """
        Переключение страниц скриптов на первую
        """
        c = Ready
        
        c.page = 1
        c.page_list = 0

        c.center_btn.configure(text = f"{c.page}/{c.len_ready_text}")
        
        cur.execute(f"""
                    SELECT ready_text.title, ready_text.text
                    FROM ready_text
                    JOIN ready_category ON ready_category.id = ready_text.id_category
                    WHERE ready_category.title = "{c.category}"
                    ORDER BY ready_text.id
                    LIMIT 10 OFFSET {c.page_list};
                    """)
        val_list = cur.fetchall() # Новый список скриптов

        for i_num, i_val in enumerate(c.btn_script_ready_list): # Переименование текста и закрепленных за кнопкой скриптов
            try:
                key = val_list[i_num][0] # Титульник скрипта
                val = val_list[i_num][1] # Текст скрипта

                i_val.configure(text = key, state='normal', bg = color["Активная"], command = lambda val=val: copy(val))

            except IndexError:
                i_val.configure(text = "", state='disabled', bg = color["Неактивная"])


    def ready_swap_category(category):
        """
        Переключение категорий скриптов для вкладки готовых скриптов
        """
        c = Ready
        c.category = category
        c.page = 1
        c.page_list = 0

        cur.execute(f"""
        SELECT ready_text.title, ready_text.text 
        FROM ready_text
        JOIN ready_category ON 
        ready_category.id = ready_text.id_category
        WHERE ready_category.title = "{c.category}"
        ORDER BY ready_text.id;
        """)
        ready_text = cur.fetchall() # Новый список скриптов
        
        len_ready_text = math.ceil(len(ready_text) / 10) # новое максимальное значение для колличества страниц
        if len_ready_text == 0:
            len_ready_text = 1
            
        c.len_ready_text = len_ready_text

        for i_val in c.category_list[:-1]: # Перекраска кнопок категорий.
            if i_val["text"] == category:
                i_val.configure(state='disabled', bg = color["Активная категория"]) 
            else:
                i_val.configure(state='normal', bg = color["Активная"])

        for i_val in range(10): # Переименование кнопок из ready_text 
            try:
                key = ready_text[i_val][0] # Титульник скрипта
                val = ready_text[i_val][1] # Текст скрипта

                c.btn_script_ready_list[i_val].configure(text = key, state='normal', bg = color["Активная"], command= lambda val=val: copy(val))

            except IndexError: # Если скрипты закончились в списке то кнопка неактивна
                c.btn_script_ready_list[i_val].configure(text = "", state='disabled', bg = color["Неактивная"]) 
        
        c.center_btn.configure(text = f"1/{len_ready_text}")

    def scroll(event):
        Ready.Canvas_top.xview_scroll(int(-1*(event.delta/120)), "units")

    Canvas_top = Canvas(main.frame_Ready, height=25) # Фрейм для кнопок категорий и скроллбара
    Canvas_top.pack(anchor=NW, fill=X)

    Сanvas_frame = Frame(Canvas_top, height=25)  # Фрейм для кнопок категорий

    Canvas_top_scrollbar = Frame(main.frame_Ready, height=10) # Фрейм для скроллбара
    Canvas_top_scrollbar.pack(anchor=NW, fill=X)

    scrollbar = Scrollbar(Canvas_top_scrollbar, orient=HORIZONTAL, command=Canvas_top.xview) # скроллбар для прокрутки 
    scrollbar.pack(expand=True, fill=BOTH)
    
    Frame_bottom = Frame(main.frame_Ready) # Нижний фрейм для стрелок и кнопок
    Frame_bottom.pack(expand=True, fill=BOTH)

    Frame_bottom_arrow = Frame(Frame_bottom) # Фрейм для стрелок
    Frame_bottom_arrow.pack(anchor=NW, fill=X)
    
    Frame_bottom_text = Frame(Frame_bottom) # Фрейм для кнопок
    Frame_bottom_text.pack(anchor=NW, expand=True, fill=BOTH)
    
    cur.execute("""SELECT title FROM ready_category ORDER BY num;""")
    ready_category = cur.fetchall() # список категорий

    category = ready_category[0][0] # Титульник первой категории в спике категорий
    
    cur.execute(f"""
    SELECT ready_text.title, ready_text.text 
    FROM ready_text
    JOIN ready_category ON 
    ready_category.id = ready_text.id_category
    WHERE ready_category.num = 1
    ORDER BY ready_text.id;
    """)
    ready_text = cur.fetchall() # Список всех значений с выбранной категории

    with_list = int() # Размер Canvas_top

    page = 1 # Страница скриптов
    page_list = 0 # Начальный элемент страницы скриптов
    
    len_ready_text = math.ceil(len(ready_text) / 10) # Максимальное колличество страниц
    if len_ready_text == 0:
        len_ready_text = 1
    
    category_list = list() # Лист для категорий скриптов
    btn_script_ready_list = list() # Лист для кнопок категорий скриптов

    for i_num, i_val in enumerate(ready_category): # Задать название кнопкам категорий btn_readys из списка категорий ready_category
        
        val = i_val[0] # Титульник категорий
        
        width_btn = int(len(str(val))) # длинна кнопки
        with_list += width_btn

        #  Кнопка категории
        btn_readys = Button(Сanvas_frame, text = val, width = width_btn,  bg = color["Активная"], font = font["Большой"], command = lambda val=val: Ready.ready_swap_category(val)) 
        btn_readys.pack(side=LEFT, expand=True, fill="both")
        category_list.append(btn_readys)

    category_list[0].configure(state='disabled', bg = color["Активная категория"])
    btn_readys_end = Button(Сanvas_frame, text = "   " * 100, state='disabled', bg = color["Активная"]) 
    btn_readys_end.pack(side=LEFT, expand=True, fill="both")
    category_list.append(btn_readys_end)

    left_btn = Button(Frame_bottom_arrow, text = "<<<<<", bg = color["Активная"], font = font["Средний"], command = ready_back_page) # Кнопка страница скриптов назад
    left_btn.pack(side = LEFT, expand=True, fill="both")

    center_btn = Button(Frame_bottom_arrow, text = f"{page}/{len_ready_text}", width = 4, bg = color["Активная"], font = font["Средний"], command = ready_first_page) # Кнопка первая страница скриптов 
    center_btn.pack(side = LEFT)
    
    right_btn = Button(Frame_bottom_arrow, text = ">>>>>", bg = color["Активная"], font = font["Средний"], command = ready_next_page) # Кнопка страница скриптов вперед
    right_btn.pack(side = LEFT, expand=True, fill="both")

    for i_val in range(10): # Создание 10 кнопок и генерация названий кнопок 
        
        try:
            key = ready_text[i_val][0] # Титульник кнопки
            val = ready_text[i_val][1] # Текст кнопки

            btn_script_readys = Button(Frame_bottom_text, text = key, bg = color["Активная"], font = font["Большой"], command= lambda val=val: copy(val)) # Кнопка скрипта категори

        except IndexError: # Если в списке закончились названия кнопок то кнопка неактивна
            btn_script_readys = Button(Frame_bottom_text, text = "", state='disabled', bg = color["Неактивная"], font = font["Большой"]) 
        finally:
            btn_script_readys.pack(fill = X, pady = 10)
            btn_script_ready_list.append(btn_script_readys)

    Canvas_top.create_window(0, 0, anchor=NW, window=Сanvas_frame, height=25) # Создание холста на Canаvas_top
    Canvas_top.config(scrollregion=(0, 0, with_list*12, 0)) # Определение длинны скрола Canvas_top
    Canvas_top.bind_all("<MouseWheel>", scroll)
    Canvas_top.config(xscrollcommand=scrollbar.set)
class Assembling:
    """
    Вкладка сборочных скриптов.

    /////// Функции /////// 
    add_copy - Функция кнопки добавления скрипта

    assembling_next_page, assembling_back_page, assembling_first_page - Управление страницами скриптов

    """

    
    def add_copy(val, buffer):
        """
        Добавление к копированным данным в буфер обмена

        val - Копируемый текст
        buffer - Текстовое поле где отражен скопированный текст
        """
            
        text = pyperclip.paste() + val.format(main.aut_text_entry.get()) # Соединяет содержание буфера обмена с val. 

        pyperclip.copy(text) # копирует text в буфер обмена

        buffer.insert(END, val.format(main.aut_text_entry.get())) # Вставляет в текстовое поле val


    def assembling_next_page():
        """
        Переключение страниц скриптов вперед для вкладки сборочных скриптов
        """
        c = Assembling
        if c.page != c.len_assembling_list:
            c.page += 1
            c.page_list += 14
            c.center_btn.configure(text = f"{c.page}/{c.len_assembling_list}")

            cur.execute(f"SELECT title, text FROM assembling_list LIMIT 13 OFFSET {c.page_list - 1};")
            val_list = cur.fetchall() # Новый список скриптов

            for i_num, i_val in enumerate(c.label_list): # Переименование текста скриптов
                try:
                    i_val.configure(text = val_list[i_num][0])
                    
                except IndexError: # Если скрипты закончились в списке то текста нет
                    i_val.configure(text = "")
            
            for i_num, i_val in enumerate(c.btn_assembling_list): # Переименование закрепленных скриптов на кнопках копирования
                try:
                    val = val_list[i_num][1]
                    i_val.configure(bg = color["Активная"], command = lambda val=val: copy(val, c.buffer_assembling))

                except IndexError: # Если скрипты закончились в списке то кнопки неактивны
                    i_val.configure(state='disabled', bg = color["Неактивная"])

            for i_num, i_val in enumerate(c.add_btn_assembling_list): # Переименование закрепленных скриптов на кнопках добавления
                try:
                    val = val_list[i_num][1]
                    i_val.configure(bg = color["Активная"], command = lambda val=val: c.add_copy(val, c.buffer_assembling))

                except IndexError: # Если скрипты закончились в списке то кнопки неактивны
                    i_val.configure(state='disabled', bg = color["Неактивная"])
        

    def assembling_back_page():
        """
        Переключение страниц скриптов назад для вкладки сборочных скриптов
        """
        c = Assembling
        if c.page != 1:
            c.page -= 1
            c.page_list -= 14
            c.center_btn.configure(text = f"{c.page}/{c.len_assembling_list}")

            cur.execute(f"SELECT title, text FROM assembling_list LIMIT 13 OFFSET {c.page_list - 1};")
            val_list = cur.fetchall() # Новый список скриптов

            for i_num, i_val in enumerate(c.label_list): # Переименование текста скриптов
                try:
                    val = val_list[i_num][0] # Значение титульника
                    i_val.configure(text = val)
                except IndexError: # Если скрипты закончились в списке то текста нет
                    i_val.configure(text='')

            for i_num, i_val in enumerate(c.btn_assembling_list): # Переименование закрепленных скриптов на кнопках копирования
                try:
                    val = val_list[i_num][1] # Значение текста скрипта
                    i_val.configure(state='normal', bg = color["Активная"], command = lambda val=val: copy(val, c.buffer_assembling))
                except IndexError: # Если скрипты закончились в списке то кнопки неактивны
                    i_val.configure(state='disabled', bg = color["Неактивная"])

            for i_num, i_val in enumerate(c.add_btn_assembling_list): # Переименование закрепленных скриптов на кнопках добавления
                try:
                    val = val_list[i_num][1] # Значение текста скрипта
                    i_val.configure(state='normal', bg = color["Активная"], command = lambda val=val: c.add_copy(val, c.buffer_assembling))
                except IndexError: # Если скрипты закончились в списке то кнопки неактивны
                    i_val.configure(state='disabled', bg = color["Неактивная"])

    def assembling_first_page():
        """
        Переключение страниц скриптов на первую для вкладки сборочных скриптов
        """
        c = Assembling
        c.page = 1
        c.page_list = 0
        c.center_btn.configure(text = f"{c.page}/{c.len_assembling_list}")

        cur.execute(f"SELECT title, text FROM assembling_list LIMIT 13 OFFSET {c.page_list - 1};")
        val_list = cur.fetchall() # Новый список скриптов

        for i_num, i_val in enumerate(c.label_list): # Переименование текста скриптов
            try:
                val = val_list[i_num][0] # Значение титульника
                i_val.configure(text = val)
            except IndexError:
                i_val.configure(text = "")
        
        for i_num, i_val in enumerate(c.btn_assembling_list): # Переименование закрепленных скриптов на кнопках копирования
            try:
                val = val_list[i_num][1] # Значение текста скрипта
                i_val.configure(state='normal', bg = color["Активная"], command = lambda val=val: copy(val, c.buffer_assembling))
            except IndexError:
                i_val.configure(state='disabled', bg = color["Неактивная"])

        for i_num, i_val in enumerate(c.add_btn_assembling_list): # Переименование закрепленных скриптов на кнопках добавления
            try:
                val = val_list[i_num][1] # Значение текста скрипта
                i_val.configure(state='normal', bg = color["Активная"], command = lambda val=val: c.add_copy(val, c.buffer_assembling))
            except IndexError:
                i_val.configure(state='disabled', bg = color["Неактивная"])
    """ 
    
    /////// Элементы /////// 

    btn_assembling - 
    add_btn_assembling - 
    lb - 


    """
    frame_assembling_bot_arrow = Frame(main.frame_assembling, height=25) # Фрейм для кнопок переключения
    frame_assembling_bot_arrow.pack(fill = X)

    frame_assembling_script = Frame(main.frame_assembling, height=400) # Фрейм для кнопок Скриптов
    frame_assembling_script.pack(fill = X)

    frame_assembling_bot_text = Frame(main.frame_assembling, height=300) # Фрейм для текстового поля буфера
    frame_assembling_bot_text.pack(fill = X)


    page = 1 # Страница скриптов
    page_list = 0 # Номер начального скрипта  

    cur.execute("SELECT title, text FROM assembling_list;")
    assembling_list = cur.fetchall() # список титульников и текстов сборочных скриптов

    len_assembling_list = math.ceil(len(assembling_list) / 14)
    if len_assembling_list == 0:
        len_assembling_list = 1
        
    btn_assembling_list = [] # Список кнопок для копирования
    add_btn_assembling_list = [] # Список кнопок для добавления к скопированному
    label_list = [] # Список текстов

    left_btn = Button(frame_assembling_bot_arrow, text = "<<<<<", bg = color["Активная"], font = font["Средний"], command = assembling_back_page) # Кнопка страница скриптов назад
    left_btn.pack(side=LEFT, expand=True, fill="both")

    center_btn = Button(frame_assembling_bot_arrow, text = f"{page}/{len_assembling_list}", width = 5,  bg = color["Активная"], font = font["Средний"], command = assembling_first_page) # Кнопка первая страница скриптов 
    center_btn.pack(side=LEFT)

    right_btn = Button(frame_assembling_bot_arrow, text = ">>>>>", bg = color["Активная"], font = font["Средний"], command = assembling_next_page) # Кнопка страница скриптов вперед
    right_btn.pack(side=LEFT, expand=True, fill="both")

    for i_val in range(13): # Создает 14 фреймов для кнопко и присваиваем им текст и титульники
        
        frame_assembling_script_panel = Frame(frame_assembling_script, height=30, highlightthickness = 1, highlightbackground=color["Обводка_сборочных"], relief=SOLID)
        frame_assembling_script_panel.pack(fill = X)

        frame_assembling_script_panel_btn = Frame(frame_assembling_script_panel, height=30)
        frame_assembling_script_panel_btn.pack(side=LEFT)

        frame_assembling_script_panel_script = Frame(frame_assembling_script_panel, height=30)
        frame_assembling_script_panel_script.pack(side=LEFT)

        try:
            key = assembling_list[i_val][0] # титульник кнопки
            val = assembling_list[i_val][1] # Текст кнопки

            btn_assembling = Button(frame_assembling_script_panel_btn, text = "С", width = 3, bg = color["Активная"], font = font["Большой"], command = lambda val=val: copy(val, Assembling.buffer_assembling)) # Кнопка копировать
            add_btn_assembling = Button(frame_assembling_script_panel_btn, text = "+", width = 3, bg = color["Активная"], font = font["Большой"], command = lambda val=val: Assembling.add_copy(val, Assembling.buffer_assembling)) # кнопка добавить к скопированному
            lb = Label(frame_assembling_script_panel_script,  text = key, font = font["Большой"]) # Текстовое поле для титульника кнопки

        except IndexError:

            btn_assembling = Button(frame_assembling_script_panel_btn, text = "С", width = 3, state='disabled', font = font["Большой"], bg = color["Неактивная"]) # Кнопка копировать неактивная
            add_btn_assembling = Button(frame_assembling_script_panel_btn, text = "+", width = 3, state='disabled', font = font["Большой"], bg = color["Неактивная"]) # кнопка добавить к скопированному неактивная
            lb = Label(frame_assembling_script_panel_script,  text = "", font = font["Большой"]) # Текстовое поле для титульника кнопки неактивная
           
        finally:
            btn_assembling.pack(side=LEFT)
            btn_assembling_list.append(btn_assembling)

            add_btn_assembling.pack(side=LEFT)
            add_btn_assembling_list.append(add_btn_assembling)

            lb.grid(column = 0, row = 0)
            label_list.append(lb)


    buffer_assembling = Text(frame_assembling_bot_text, font = font["Большой"],wrap=WORD) # текстовое поле для отображения скопированного
    buffer_assembling.pack(fill=BOTH)
    buffer_assembling.bind("<Key>", lambda e: "break") 

class Settings:
    """
    Вкладка настроек
    """
    notebook = ttk.Notebook(main.frame_settings)
    notebook.pack(expand=True, fill=BOTH)

    # создаем пару фреймвов
    frame_assembling = ttk.Frame(notebook) # Настройка сборочных скриптов
    frame_Ready = ttk.Frame(notebook) # Настройка Готовых скриптов
    frame_general = ttk.Frame(notebook) # Общие настройки

    frame_assembling.pack(fill=BOTH, expand=True)
    frame_Ready.pack(fill=BOTH, expand=True)
    frame_general.pack(fill=BOTH, expand=True)


    # добавляем фреймы в качестве вкладок
    notebook.add(frame_assembling, text="Сборка")
    notebook.add(frame_Ready, text="Готовые")
    notebook.add(frame_general, text="Общие")


class Settings_Assembling:
    """
    Вкладка настроек сборочных скриптов

    /////// Функции /////// 

    commit_script - перекрашивает экран после удаления, сохранения , изменения
    switching_assembling_combobox - Евент при переключении значений комбобокса choice_val_menu
    assembling_delete - Удаляет выбранный скрипт
    assembling_save - Добавляет новый скрипт или изменяет старый
    """
    def commit_script(status):
        c = Settings_Assembling

        conn.commit()
        cur.execute("SELECT id, title, text FROM assembling_list;") 
        c.val_list = cur.fetchall() # список сборочных скриптов с айди титульником и текстом 

        num_list = [i_val[0] for i_val in c.val_list] 
        num_list.append(len(num_list) + 1)
        c.num_list = num_list # новый список номеров скриптов

        title_list = [f"{i_val[0]}| " + i_val[1] for i_val in c.val_list]
        c.title_list = ["+ Добавить новый скрипт"] + title_list + ["+ Добавить новый скрипт"] # новый список скриптов
        
        c.choice_variable_num.set(1) # изменение номера скрипта на 1
        c.choice_variable_val.set(c.title_list[0]) # изменение текста скрипта на первый + Добавить новый скрипт
        c.choice_num_menu.configure(values = c.num_list) # Изменение списка номеров
        c.choice_val_menu.configure(values = c.title_list) # Изменение списка скриптов
        
        c.title_empty.delete(0, END)
        c.text_empty.delete("1.0", END)

        c.btn_save.configure(bg = color["Сохранить"], text = "Сохранить")
        c.btn_delete.configure(state= "disable")

        if status == "Удалить":         
            c.title_empty.insert(0, "Удалено")

            c.title_empty.configure(bg = color["Удалить"])
            c.text_empty.configure(bg = color["Удалить"])

        elif status == "Сохранить":
            c.title_empty.configure(bg = color["Сохранить"])
            c.text_empty.configure(bg = color["Сохранить"])

            c.title_empty.insert(0, "Сохранено")

        elif status == "Изменить":
            c.title_empty.configure(bg = color["Изменить"])
            c.text_empty.configure(bg = color["Изменить"]) 

            c.title_empty.insert(0, "Изменено")

    def switching_assembling_combobox(event):
        """
        Изменяет значения полей title_empty, text_empty, choice_num_menu, btn_save, btn_delete при переключении комбобокса choice_val_menu
        """
        c = Settings_Assembling
        choice = c.choice_val_menu.get() # выбранный элемент в комбобоксе

        if choice == "+ Добавить новый скрипт":
            c.title_empty.delete(0, END)
            c.title_empty.insert(0, "Титульник")

            c.text_empty.delete("1.0", END)
            c.text_empty.insert("1.0", "Текст скрипта\n\n{} - Символ для автоматической вставки")

            c.choice_num_menu.configure(values = c.num_list)

            c.btn_save.configure(text = "Сохранить", bg = color["Сохранить"])
            c.btn_delete.configure(state='disabled')
        else:
            index_title = int(choice[:choice.index('|')]) # индекс титульника выбранного скрипта
            
            text = c.val_list[index_title - 1][2] # текст выбранного скрипта
            c.choice_variable_num.set(index_title) 

            c.title_empty.delete(0, END)
            c.title_empty.insert(0, choice[choice.index(" ") + 1:])

            c.text_empty.delete("1.0", END)
            c.text_empty.insert("1.0", text)
            
            num_list = c.num_list[:len(c.num_list) - 1] # Если выбираем действующий скрипт то лист номеров на 1 меньше
            c.choice_num_menu.configure(values = num_list)

            c.btn_save.configure(text = "Изменить", state='normal', bg = color["Изменить"])
            c.btn_delete.configure(state='normal')
        
        c.title_empty.configure(bg = color["Текст"])
        c.text_empty.configure(bg = color["Текст"]) 



    def assembling_delete():
        """
        Удаляет выбранный скрипт
        """
        c = Settings_Assembling
        choice = c.choice_val_menu.get() # Значение выбранного элемента вв комбобоксе
        index_title = int(choice[:choice.index('|')]) #  Порядковый номер титульника (14| Титульник) index_title = 14
        title = choice[choice.index(' ') + 1:]
        select_script = cur.execute(f"SELECT id, title FROM assembling_list WHERE id = '{index_title}';").fetchall()[0] # ищем скрипт в базе данных по id

        if select_script == (index_title, title):
            cur.execute(f"DELETE FROM assembling_list WHERE id = '{index_title}';") # Удаляет скрипт с id = index_title
            cur.execute(f"UPDATE assembling_list SET id = id - 1 WHERE id > '{index_title}'") # Перемещяет все скрипты выше index_title на 1 id ниже
            c.commit_script("Удалить")
        else:
            error_two_open()
    

    def assembling_save():
        """
        Добавляет или изменяет выбранный скрипт
        """
        c = Settings_Assembling
        choice = choice = c.choice_val_menu.get() # Значение выбранного элемента в комбобоксе
        index = int(c.choice_variable_num.get()) # Значение выбранного номера элемента в комбобоксе
        title = c.title_empty.get() # Значение поля ввода титульника 
        text = c.text_empty.get("1.0", END) # Значение поля ввода текста 

        if choice == "+ Добавить новый скрипт":
            max_index = len(c.val_list) + 1 # Максимальное значение номера скрипта + 1

            for i_index in range(max_index, index - 1, -1):  # Добавляет 1 к каждому скрипту чей id = i_index    
                cur.execute(f"UPDATE assembling_list SET id = id + 1 WHERE id = '{i_index}'") # Увеличивает значение id у скрипта где id = i_index
                conn.commit()

            cur.execute("INSERT INTO assembling_list(id, title, text) VALUES (?, ?, ?)", (index, title, text[:-1])) # Добавляет скрипт в БД
            conn.commit()

            c.commit_script("Сохранить")

        else:
            index_title = int(choice[:choice.index('|')]) # Порядковый номер титульника (14| Титульник) index_title = 14

            title_script = choice[choice.index(' ') + 1:]
            select_script = cur.execute(f"SELECT id, title FROM assembling_list where id = '{index_title}';").fetchall()[0]
            
            if select_script == (index_title, title_script):
                if index_title == index:
                    cur.execute(f"UPDATE assembling_list SET title = ?, text = ? WHERE id = ?", (title, text[:-1], index_title)) # Изменяет текст и титульник скрипта с id = index_title
                    
                elif index_title < index:    
                    cur.execute(f"DELETE FROM assembling_list where id = '{index_title}';") # Удаляет скрипта с id = index_title
                    cur.execute(f"UPDATE assembling_list SET id = id - 1 WHERE id > '{index_title}' AND id <= '{index}'") # Изменяет id скриптов у которых id больше index_title и меньше index_title + 1
                    cur.execute("INSERT INTO assembling_list(id, title, text) VALUES (?, ?, ?)", (index, title, text[:-1])) # Добавляет скрипт в БД
                    

                else:
                    cur.execute(f"DELETE FROM assembling_list where id = '{index_title}';") # Удаляет скрипта с id = index_title
                    
                    for i_index in range(index_title - 1, index - 1, -1): # Добавляет +1 id всем элементам от нового до старого индекса
                        cur.execute(f"UPDATE assembling_list SET id = id + 1 WHERE id = '{i_index}'") # Увеличивает значение id у скрипта где id = i_index
                        conn.commit()

                    cur.execute("INSERT INTO assembling_list(id, title, text) VALUES (?, ?, ?)", (index, title, text[:-1])) # Добавляет скрипт в БД
                    
                c.commit_script("Изменить")
            else:
                error_two_open()


    frame_choice = Frame(Settings.frame_assembling) # Фрейм для выбора названия скрипта
    frame_choice.pack(fill=X, pady = 5)

    frame_title = Frame(Settings.frame_assembling) # Фрейм для имени скрипта
    frame_title.pack(fill=X, pady = 5)

    frame_text = Frame(Settings.frame_assembling) # Фрейм для текста скрипта
    frame_text.pack(fill=X, pady = 5)

    frame_btn = Frame(Settings.frame_assembling) # Фрейм для кнопок удалить сохранить изменить
    frame_btn.pack(fill=BOTH)

    frame_choice_num = Frame(frame_choice) # Дочерний фрейм frame_choice для нумерации скрипта
    frame_choice_num.pack(anchor=NW, side=LEFT, fill=X)

    frame_choice_val = Frame(frame_choice) # Дочерний фрейм frame_choice для названий скриптов
    frame_choice_val.pack(anchor=NW, side=LEFT, fill=X)

    cur.execute("SELECT id, title, text FROM assembling_list;")
    val_list = cur.fetchall() # Список скриптов

    num_list = [i_val[0] for i_val in val_list]
    num_list.append(len(num_list) + 1) # Список номеров скриптов
    title_list = [f"{i_val[0]}| " + i_val[1] for i_val in val_list]
    title_list = ["+ Добавить новый скрипт"] + title_list + ["+ Добавить новый скрипт"] # Список титульников скриптов

    choice_variable_num = StringVar(frame_choice_num) # Значение в комбобоксе choice_num_menu
    choice_variable_num.set(1) 
    choice_num_menu = ttk.Combobox(frame_choice_num, width=3, state = "readonly", font = font["Средний"], textvariable=choice_variable_num, values=num_list) # комбобокс номеров скриптов
    choice_num_menu.pack()

    choice_variable_val = StringVar(frame_choice_val) # Значение в комбобоксе choice_val_menu
    choice_variable_val.set(title_list[0]) 
    choice_val_menu = ttk.Combobox(frame_choice_val, width=100, state = "readonly", font = font["Средний"], textvariable=choice_variable_val, values=title_list) # комбобокс значений скриптов
    choice_val_menu.pack()
    choice_val_menu.bind("<<ComboboxSelected>>", switching_assembling_combobox)

    title_lb = Label(frame_title, text = "Название") # Надпись для ориентации
    title_lb.pack(fill=X)
    title_empty = Entry(frame_title, font = font["Большой"]) # Поле ввода титульника
    title_empty.pack(fill=X)
    title_empty.bind(ctrl, keys)

    text_lb = Label(frame_text, text = "Текст") # Надпись для ориентации
    text_lb.pack(fill=X)
    text_empty = Text(frame_text, wrap=WORD, undo=True, font = font["Средний"]) # поле ввода текста скрипта
    text_empty.pack(side=LEFT, expand=True, fill = BOTH)
    text_empty.bind(ctrl, keys)

    btn_delete = Button(frame_btn,text = "Удалить", state= "disable", bg = color["Удалить"], command = assembling_delete) # кнопка удаления скрипта
    btn_delete.pack(side=LEFT, expand=True, fill="both")
    
    btn_save = Button(frame_btn, text = "Сохранить", bg = color["Сохранить"], command = assembling_save) # кнопка изменения или добавления скрипта
    btn_save.pack(side=LEFT, expand=True, fill="both")


class Settings_Ready:

    def commit_script(status):
        c = Settings_Ready
        conn.commit()
        cur.execute(f"SELECT num, title FROM ready_category ORDER BY num;")
        val_list = cur.fetchall()

        category_num_list = [i_val[0] for i_val in val_list]
        category_num_list.append(len(category_num_list) + 1)
        c.category_num_list = category_num_list

        category_list = [f"{i_val[0]}| " + i_val[1] for i_val in val_list]
        category_list = ["+ Добавить новую категорию"] + category_list + ["+ Добавить новую категорию"]
        c.category_list = category_list

        c.choice_category_num_menu.configure(values=c.category_num_list)
        c.choice_category_menu.configure(values=c.category_list)

        c.choice_category_num.set("1") 
        c.choice_category_val.set(category_list[0])

        c.choice_script_num.set("") 
        c.choice_script_val.set("")

        c.choice_script_menu.configure(state='disabled')
        c.choice_script_num_menu.configure(state='disabled')
        
        c.title_empty.delete(0, END)
        c.text_empty.delete("1.0", END)

        c.btn_save.configure(bg = color["Сохранить"], text = "Сохранить")
        c.btn_delete.configure(state="disabled")

        if status == "Удалить":         
            c.title_empty.insert(0, "Удалено")

            c.title_empty.configure(bg = color["Удалить"])
            c.text_empty.configure(bg = color["Удалить"])

        elif status == "Сохранить":
            c.title_empty.configure(bg = color["Сохранить"])
            c.text_empty.configure(bg = color["Сохранить"])

            c.title_empty.insert(0, "Сохранено")

        elif status == "Изменить":
            c.title_empty.configure(bg = color["Изменить"])
            c.text_empty.configure(bg = color["Изменить"]) 

            c.title_empty.insert(0, "Изменено")

        elif status == "Ошибка создания":
            c.title_empty.insert(0, "Имя занято")
            c.btn_save.configure(state="disabled")
            c.title_empty.configure(bg = color["Удалить"])
            c.text_empty.configure(bg = color["Удалить"])


    def new_category_btn():
        c = Ready
        cur.execute("""SELECT title FROM ready_category ORDER BY num;""")
        ready_category = cur.fetchall()

        for i_val in c.category_list:
            i_val.destroy()

        c.category_list.clear()
        c.with_list = 0
        for i_val in ready_category: # Цикл 4
        
            val = i_val[0]
            
            width_btn = int(len(str(val)))
            c.with_list += width_btn

            btn_readys = Button(c.Сanvas_frame, text = val, width = width_btn, font = font["Большой"],  bg = color["Активная"], command = lambda val=val: Ready.ready_swap_category(val)) 
            btn_readys.pack(side=LEFT, expand=True, fill="both")
            c.category_list.append(btn_readys)

        c.category_list[0].configure(state='disabled', bg = color["Активная категория"])

        btn_readys_end = Button(c.Сanvas_frame, text = "   " * 100, state='disabled', bg = color["Активная"]) 
        btn_readys_end.pack(side=LEFT, expand=True, fill="both")
        c.category_list.append(btn_readys_end)
        c.Canvas_top.config(scrollregion=(0, 0, c.with_list*12, 0))

    def switching_category_combobox(event):
        c = Settings_Ready
        choice = c.choice_category_val.get()

        if choice == "+ Добавить новую категорию":
            c.title_empty.delete(0, END)
            c.title_empty.insert(0, "Титульник")

            c.choice_script_num_menu.configure(state='disabled')
            c.choice_script_menu.configure(state='disabled')

            c.choice_category_num_menu.configure(values = c.category_num_list)
            
            c.btn_save.configure(text = "Сохранить", bg = color["Сохранить"], state="normal")
            c.btn_delete.configure(state='disabled')
        else:
            try:
                index_title = int(choice[:choice.index('|')])
                cur.execute(f"SELECT id FROM ready_category WHERE num = '{index_title}';")
                id_category = cur.fetchone()[0]

                cur.execute(f"SELECT id, title, text FROM ready_text WHERE id_category = {id_category} ORDER BY id;")
                c.script_list = cur.fetchall()

                c.script_num_list = [i_val[0] for i_val in c.script_list]
                c.script_num_list.append(len(c.script_num_list) + 1)
                c.script_title_list = [f"{i_val[0]}| " + i_val[1] for i_val in c.script_list]
                c.script_title_list = ["+ Добавить новый скрипт", ""] + c.script_title_list + ["", "+ Добавить новый скрипт"]
                
                category_num_list = c.category_num_list[:len(c.category_num_list)-1]
                c.choice_category_num_menu.configure(values = category_num_list)

                c.choice_script_num_menu.configure(values = c.script_num_list)
                c.choice_script_menu.configure(values = c.script_title_list , state='readonly')

                c.choice_category_num.set(index_title) 
        
                c.title_empty.delete(0, END)
                c.title_empty.insert(0, choice[choice.index(" ") + 1:])

                c.btn_save.configure(text = "Изменить", state='normal', bg = color["Изменить"])
                c.btn_delete.configure(state='normal')
            except TypeError:
                error_two_open()
        
        c.choice_script_num.set("") 
        c.choice_script_val.set("") 
        c.choice_script_num_menu.configure(state='disabled')
        c.text_empty.delete("1.0", END)
        c.text_empty.configure(state="disabled")
        
        c.title_empty.configure(bg = color["Текст"])
        c.text_empty.configure(bg = color["Текст"])


    def ready_delete():
        c = Settings_Ready

        choice_category = c.choice_category_val.get()
        title_category = choice_category[choice_category.index(' ') + 1:]
        choice_script = c.choice_script_val.get()
        index_category = int(choice_category[:choice_category.index('|')])
        select_category = cur.execute(f"SELECT title FROM ready_category WHERE num = '{index_category}';").fetchall()
        try:
            if choice_script == "" and select_category[0][0] == title_category:
                    conn.execute("PRAGMA foreign_keys = 1")
                    cur.execute(f"DELETE FROM ready_category WHERE num = '{index_category}';")
                    cur.execute(f"UPDATE ready_category SET num = num - 1 WHERE num > '{index_category}'")
                    conn.commit()
                    c.new_category_btn()
                    c.commit_script("Удалить")   


            elif choice_script != "":
                index_script = int(choice_script[:choice_script.index('|')])    
                id_category =  cur.execute(f"SELECT id FROM ready_category WHERE num = '{index_category}'").fetchone()[0]

                select_script = cur.execute(f"SELECT id, title FROM ready_text WHERE id_category = '{id_category}' AND id = {index_script};").fetchall()[0]
                num_script = choice_script[choice_script.index(' ') + 1:]

                if select_script == (index_script, num_script) and select_category[0][0] == title_category:
                    cur.execute(f"DELETE FROM ready_text WHERE id_category = '{id_category}' AND id = '{index_script}';")
                    cur.execute(f"UPDATE ready_text SET id = id - 1 WHERE id_category = '{id_category}' AND id > '{index_script}'")
                    conn.commit()    
                    c.commit_script("Удалить")   
                else:
                    error_two_open()

            else:
                error_two_open()
        except IndexError:
            error_two_open()


    def ready_save():
        c = Settings_Ready
        choice_category = c.choice_category_val.get()
        choice_script = c.choice_script_val.get()
        title = c.title_empty.get()
        text = c.text_empty.get("1.0", END)
        category_num = int(c.choice_category_num.get())
        script_num = c.choice_script_num.get()
        
        cur.execute("SELECT title FROM ready_category WHERE title = ?;", (title,))
        category_choise = cur.fetchall()

        if choice_category == "+ Добавить новую категорию":
            if len(category_choise) == 0:
                cur.execute(f"UPDATE ready_category SET num = num + 1 WHERE num >= '{category_num}'")
                cur.execute("INSERT INTO ready_category(num, title) VALUES (?, ?)", (category_num, title))
                
                c.commit_script("Сохранить")
                c.new_category_btn()
            else:
                c.commit_script("Ошибка создания")
        
        elif choice_script == "":
            index_category = int(choice_category[:choice_category.index('|')])
            title_val = choice_category[choice_category.index(" ") + 1:]
            select_category = cur.execute(f"SELECT num, title FROM ready_category WHERE num = '{index_category}';").fetchall()
            
            cur.execute("SELECT title FROM ready_category WHERE title != ?;", (title_val,))
            category_list = cur.fetchall()

            if (title,) not in category_list:

                if index_category == category_num:
                    cur.execute(f"UPDATE ready_category SET title = ? WHERE num = ?", (title, category_num))
                    
                elif index_category > category_num:
                    cur.execute(f"UPDATE ready_category SET num = num + 1 WHERE num >= '{category_num}' AND num < '{index_category}'")
                    cur.execute(f"UPDATE ready_category SET num = ?, title = ? WHERE title = ?", (category_num, title, title_val))
                    
                elif index_category < category_num:
                    cur.execute(f"UPDATE ready_category SET num = num - 1 WHERE num <= '{category_num}' AND num > '{index_category}'")
                    cur.execute(f"UPDATE ready_category SET num = ?, title = ? WHERE title = ?", (category_num, title, title_val))  
                
                c.commit_script("Изменить")
                c.new_category_btn()
            elif select_category[0] != (index_category, choice_category[choice_category.index(' ') + 1:]):
                error_two_open()
            else:
                c.commit_script("Ошибка создания")
            

        elif choice_script == "+ Добавить новый скрипт":
                index_category = int(choice_category[:choice_category.index('|')])    
                id_category =  cur.execute(f"SELECT id FROM ready_category WHERE num = '{index_category}'").fetchone()[0]

                index_category = int(choice_category[:choice_category.index('|')])
                cur.execute(f"SELECT id FROM ready_category WHERE num = '{index_category}'")
                id_category = cur.fetchone()[0]

                cur.execute(f"UPDATE ready_text SET id = id + 1 WHERE id >= '{script_num}' AND id_category = {id_category}")
                cur.execute("INSERT INTO ready_text(id, id_category, title, text) VALUES (?, ?, ?, ?)", (script_num, id_category, title, text[:-1]))
                conn.commit()

                c.commit_script("Сохранить")


        elif choice_script != "+ Добавить новый скрипт":
            index_category = int(choice_category[:choice_category.index('|')])
            index_script = int(choice_script[:choice_script.index('|')])
            
            script_num = int(script_num)
            id_category =  cur.execute(f"SELECT id FROM ready_category WHERE num = '{index_category}'").fetchone()[0]

            select_script = cur.execute(f"SELECT id, title FROM ready_text WHERE id_category = '{id_category}' AND id = {index_script};").fetchall()[0]
            num_script = choice_script[choice_script.index(' ') + 1:]

            if select_script == (index_script, num_script):
                if script_num == index_script:
                    cur.execute(f"UPDATE ready_text SET title = ?, text = ? WHERE id = ? AND id_category = {id_category}", (title, text[:-1], script_num))
                    
                elif script_num > index_script:
                    cur.execute(f"DELETE FROM ready_text WHERE id = '{index_script}' AND id_category = {id_category};")
                    cur.execute(f"UPDATE ready_text SET id = id - 1 WHERE id > '{index_script}' AND id <= '{script_num}' AND id_category = {id_category}")
                    cur.execute("INSERT INTO ready_text(id, id_category, title, text) VALUES (?, ?, ?, ?)", (script_num, id_category, title, text[:-1]))
                
                elif script_num < index_script:
                    cur.execute(f"DELETE FROM ready_text WHERE id = '{index_script}' AND id_category = {id_category};")
                    cur.execute(f"UPDATE ready_text SET id = id + 1 WHERE id < '{index_script}' AND id >= '{script_num}' AND id_category = {id_category}")
                    cur.execute("INSERT INTO ready_text(id, id_category, title, text) VALUES (?, ?, ?, ?)", (script_num, id_category, title, text[:-1]))
                
                c.commit_script("Изменить")
            else:
                error_two_open()
        else:
            error_two_open()



    def switching_script_combobox(event):
        c = Settings_Ready 
        choice_script = c.choice_script_val.get()
        choice_category = c.choice_category_val.get()

        c.title_empty.delete(0, END)
        c.text_empty.delete("1.0", END)

        if choice_script == "":
            c.choice_script_num_menu.configure(state='disabled')
            c.text_empty.configure(state="disabled")
            c.choice_script_num.set("")

            c.title_empty.insert(0, choice_category[choice_category.index(" ") + 1:])
        
            c.btn_delete.configure(state="normal")
            c.btn_save.configure(text="Изменить", bg=color["Изменить"])

        elif choice_script == "+ Добавить новый скрипт":
            c.choice_script_num_menu.configure(state="readonly")
            c.choice_script_num_menu.configure(state="readonly", values = c.script_num_list)

            c.text_empty.configure(state="normal")
            if c.choice_script_num.get() == "":
                c.choice_script_num.set(1)

            c.title_empty.insert(0, "Титульник")
            c.text_empty.insert("1.0", "Текст скрипта\n\n{} - Символ для автоматической вставки")


            c.btn_delete.configure(state="disabled")
            c.btn_save.configure(text="Сохранить", bg=color["Сохранить"])
        else:
            script_num = int(choice_script[:choice_script.index('|')])

            script_num_list = c.script_num_list[:len(c.script_num_list)-1]
            c.choice_script_num_menu.configure(state="readonly", values = script_num_list)

            c.text_empty.configure(state="normal")
            c.choice_script_num.set(script_num)

            c.title_empty.insert(0, choice_script[choice_script.index(" ") + 1:])
            c.text_empty.insert("1.0", c.script_list[script_num - 1][2])

            c.btn_delete.configure(state="normal")
            c.btn_save.configure(text="Изменить", bg=color["Изменить"])

        c.title_empty.configure(bg = color["Текст"])
        c.text_empty.configure(bg = color["Текст"])

    
    frame_choice_category = Frame(Settings.frame_Ready)
    frame_choice_category.pack(fill=X)
        
    frame_choice_script = Frame(Settings.frame_Ready)
    frame_choice_script.pack(fill=X, pady = 5)

    frame_title = Frame(Settings.frame_Ready)
    frame_title.pack(fill=X, pady = 5)

    frame_text = Frame(Settings.frame_Ready)
    frame_text.pack(fill=X, pady = 5)

    frame_btn = Frame(Settings.frame_Ready)
    frame_btn.pack(fill=BOTH)

    frame_choice_category_num = Frame(frame_choice_category)
    frame_choice_category_num.pack(anchor=NW, side=LEFT, fill = X)

    frame_choice_category_val = Frame(frame_choice_category)
    frame_choice_category_val.pack(anchor=NW, side=LEFT, fill = X)
    
    frame_choice_script_num = Frame(frame_choice_script)
    frame_choice_script_num.pack(anchor=NW, side=LEFT, fill = X)

    frame_choice_script_val = Frame(frame_choice_script)
    frame_choice_script_val.pack(anchor=NW, side=LEFT, fill = X)

    """--- Раздел с категориями  ---"""

    cur.execute("SELECT num, title FROM ready_category ORDER BY num;")
    category_list = cur.fetchall()

    category_num_list = [i_val[0] for i_val in category_list]
    category_num_list.append(len(category_num_list) + 1)
    category_title_list = [f"{i_val[0]}| " + i_val[1] for i_val in category_list]
    category_title_list = ["+ Добавить новую категорию"] + category_title_list + ["+ Добавить новую категорию"]
    category_id = 1
    
    choice_category_num = StringVar(frame_choice_category_num)
    choice_category_num.set(1) 
    choice_category_num_menu = ttk.Combobox(frame_choice_category_num, width=3, state = "readonly", font = font["Средний"], textvariable=choice_category_num, values=category_num_list)
    choice_category_num_menu.pack()

    choice_category_val = StringVar(frame_choice_category_val)
    choice_category_val.set(category_title_list[0]) 
    choice_category_menu = ttk.Combobox(frame_choice_category_val, width=100, state = "readonly", font = font["Средний"], textvariable=choice_category_val, values=category_title_list)
    choice_category_menu.pack()
    choice_category_menu.bind("<<ComboboxSelected>>", switching_category_combobox)
    
    """--- Раздел с скриптами ---"""
    script_num_list = list()
    script_title_list = list()

    script_list = list()

    choice_script_num = StringVar(frame_choice_script_num)
    choice_script_num.set("") 
    choice_script_num_menu = ttk.Combobox(frame_choice_script_num, width=3, state = "disabled", textvariable=choice_script_num)
    choice_script_num_menu.pack()

    choice_script_val = StringVar(frame_choice_script_val)
    choice_script_val.set("") 
    choice_script_menu = ttk.Combobox(frame_choice_script_val, width=100, state = "disabled", textvariable=choice_script_val)
    choice_script_menu.pack()
    choice_script_menu.bind("<<ComboboxSelected>>", switching_script_combobox)

    """--- Раздел полями ввода  ---"""
    title_lb = Label(frame_title, text = "Название")
    title_lb.pack(fill=X)
    title_empty = Entry(frame_title, width=100, bg = color["Текст"], font = font["Большой"])
    title_empty.pack()
    title_empty.bind(ctrl, keys)

    text_lb = Label(frame_text, text = "Текст")
    text_lb.pack(fill=X)
    text_empty = Text(frame_text, state= "disable", bg = color["Текст"], font = font["Средний"], wrap=WORD, undo=True)
    text_empty.pack()
    text_empty.bind(ctrl, keys)

    btn_delete = Button(frame_btn, text = "Удалить", state= "disable", bg = color["Удалить"], command=ready_delete)
    btn_delete.pack(side=LEFT, expand=True, fill="both")
    
    btn_save = Button(frame_btn, text = "Сохранить", bg = color["Сохранить"], command=ready_save)
    btn_save.pack(side=LEFT, expand=True, fill="both")


class Book:
    """
    Вкладка Блокнота
    """
    def save_to_database(event=None): # Сохраняет важные записи в БД
        c = Book
        new_note = c.book_important_text.get("1.0", "end-1c")

        cur.execute("UPDATE setting SET important_book_text = ? WHERE id = 1", (new_note,))
        conn.commit()

    cur.execute("SELECT important_book, important_book_text FROM setting;")
    setting = cur.fetchall()

    book_text = Text(main.frame_book, wrap=WORD, undo=True, font = font["Большой"]) # Поле для блокнота растянутое на всю ширину вкладки
    book_text.pack(expand=True, fill = BOTH)
    book_text.bind(ctrl, keys)

    book_important_text = Text(main.frame_book, height=10, font = font["Большой"], state= "normal", wrap=WORD, undo=True) # Поле для важных записей
    book_important_text.insert("1.0", setting[0][1])
    book_important_text.bind(ctrl, keys)
    book_important_text.bind('<KeyRelease>', save_to_database)

    if setting[0][0] == 1:
        book_important_text.pack(fill = X)
    else:
        book_important_text.pack_forget()
    

class Settings_general:

    def export_file():
        c = Settings_general
        category = cur.execute("""SELECT id, title FROM ready_category ORDER BY num""").fetchall()
        assembling = cur.execute("""SELECT title, text FROM assembling_list""").fetchall()

        exel = Workbook()
        list_exel = exel.sheetnames

        exel.create_sheet("Сборочные")
        exel_assembling = exel["Сборочные"]

        for i_num, i_val in enumerate(assembling):
            title = exel_assembling.cell(row=i_num + 1, column=1)
            title.value = i_val[0]

            text = exel_assembling.cell(row=i_num + 1, column=2)
            text.value = i_val[1]

        for i_val in category:
            category_id = i_val[0]
            category_title = i_val[1]


            cur.execute(f"""SELECT title, text FROM ready_text WHERE id_category = {category_id} ORDER BY id;""")
            script = cur.fetchall()

            exel.create_sheet(str(category_title))
            exel_category = exel[category_title]

            for i_num, i_val in enumerate(script):

                title = exel_category.cell(row=i_num + 1, column=1)
                title.value = i_val[0]

                text = exel_category.cell(row=i_num + 1, column=2)
                text.value = i_val[1]

        exel.remove(exel[list_exel[0]])

        my_dir_exel = os.path.abspath(os.path.join(my_dir, os.pardir))
        print(my_dir_exel)
        exel_file = os.path.join(my_dir_exel, "Мои скрипты.xlsx") # Ссылка на иконку приложения
        try:
            if os.path.exists(exel_file): # Если сылка найдена то удалить и сохранить
                os.remove(exel_file)
            exel.save(os.path.abspath(os.path.join(my_dir_exel, 'Мои скрипты.xlsx')))
            c.entry_link.delete("1.0", END)
            c.entry_link.insert("1.0", f"Данные выгружены сюда --> {my_dir_exel}")
            c.entry_link.configure(bg = color["Сохранить"])
        except PermissionError:
            c.entry_link.delete("1.0", END)
            c.entry_link.insert("1.0", "Пожалуйста, закройте Exel файл")
            c.entry_link.configure(bg = color["Удалить"])

    def book_checkbutton_changed():
        
        cur.execute("SELECT important_book FROM setting WHERE id = 1;")
        important_book = cur.fetchone()

        if important_book[0] == 1:
            Book.book_important_text.pack_forget()
            new = 0
        else:
            Book.book_important_text.pack(fill = X)
            new = 1

        cur.execute(f"UPDATE setting SET important_book = ? WHERE id = 1;", (new,))
        conn.commit()

    if Book.setting[0][0] == 1:
        num = 1
    else: 
        num = 0

    var = IntVar(value=num)
    book_checkbutton = Checkbutton(Settings.frame_general, text="Блокнот для заметок", font = font["Большой"], variable=var, onvalue=1, offvalue=0, command=book_checkbutton_changed)
    book_checkbutton.pack(padx=6, pady=6, anchor=NW)

    link_frame = Frame(Settings.frame_general)
    link_frame.pack(fill=X, side=BOTTOM)

    entry_link = Text(link_frame, height=4, state="normal")
    entry_link.pack(fill=X)
    entry_link.bind("<Key>", lambda e: "break")

    link_frame_btn = Frame(link_frame)
    link_frame_btn.pack(fill=X)

    export_btn = Button(link_frame_btn, text = "Выгрузить", font = font["Большой"], command=export_file)
    export_btn.pack(fill=X, expand=True, side=LEFT)
    
    # import_btn = Button(link_frame_btn, text = "Загрузить", font = font["Большой"])
    # import_btn.pack(fill=X, expand=True, side=LEFT)

root.mainloop()

